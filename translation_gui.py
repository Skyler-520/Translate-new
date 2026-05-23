# -*- coding: utf-8 -*-
"""
自动化翻译处理工具 - GUI版本
功能描述:
    1. 支持文件夹和文件两种输入模式
    2. 支持超过100种语言的翻译
    3. 提供翻译确认机制和进度显示
    4. 自动生成各语言版本的XML文件
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext, simpledialog
import xml.etree.ElementTree as ET
import json
import os
import requests
from urllib.parse import quote
import time
import threading
import hashlib
import uuid
from openpyxl import Workbook, load_workbook
import argparse
import tempfile
import logging
import re

# 语言映射字典 - 包含语言代码、名称和中文说明
# 格式: {'语言缩写': {'code': 'Google翻译代码', 'name': '语言名称 (中文语言:中文国家)'}}
LANG_MAP = {
    'CAT': {'code': 'ca', 'name': 'Catalan: Spain (加泰罗尼亚语:西班牙)'},
    'CHT': {'code': 'zh-TW', 'name': 'Chinese: Traditional (Taiwan) (中文繁体:中国台湾)'},
    'CHS': {'code': 'zh-CN', 'name': 'Chinese: Simplified (PRC) (中文简体:中国)'},
    'ZHH': {'code': 'zh-HK', 'name': 'Chinese: Hong Kong S.A.R. (中文:中国香港)'},
    'ZHI': {'code': 'zh-SG', 'name': 'Chinese: Singapore (中文:新加坡)'},
    'ZHM': {'code': 'zh-MO', 'name': 'Chinese: Macau SAR (中文:中国澳门)'},
    'ARA': {'code': 'ar-SA', 'name': 'Arabic: Saudi Arabia (阿拉伯语:沙特阿拉伯)'},
    'ARI': {'code': 'ar-IQ', 'name': 'Arabic: Iraq (阿拉伯语:伊拉克)'},
    'ARE': {'code': 'ar-EG', 'name': 'Arabic: Egypt (阿拉伯语:埃及)'},
    'ARL': {'code': 'ar-LY', 'name': 'Arabic: Libya (阿拉伯语:利比亚)'},
    'ARG': {'code': 'ar-DZ', 'name': 'Arabic: Algeria (阿拉伯语:阿尔及利亚)'},
    'ARM': {'code': 'ar-MA', 'name': 'Arabic: Morocco (阿拉伯语:摩洛哥)'},
    'ART': {'code': 'ar-TN', 'name': 'Arabic: Tunisia (阿拉伯语:突尼斯)'},
    'ARO': {'code': 'ar-OM', 'name': 'Arabic: Oman (阿拉伯语:阿曼)'},
    'ARY': {'code': 'ar-YE', 'name': 'Arabic: Yemen (阿拉伯语:也门)'},
    'ARS': {'code': 'ar-SY', 'name': 'Arabic: Syria (阿拉伯语:叙利亚)'},
    'ARJ': {'code': 'ar-JO', 'name': 'Arabic: Jordan (阿拉伯语:约旦)'},
    'ARB': {'code': 'ar-LB', 'name': 'Arabic: Lebanon (阿拉伯语:黎巴嫩)'},
    'ARK': {'code': 'ar-KW', 'name': 'Arabic: Kuwait (阿拉伯语:科威特)'},
    'ARU': {'code': 'ar-AE', 'name': 'Arabic: U.A.E. (阿拉伯语:阿联酋)'},
    'ARH': {'code': 'ar-BH', 'name': 'Arabic: Bahrain (阿拉伯语:巴林)'},
    'ARQ': {'code': 'ar-QA', 'name': 'Arabic: Qatar (阿拉伯语:卡塔尔)'},
    'BGR': {'code': 'bg', 'name': 'Bulgarian: Bulgaria (保加利亚语:保加利亚)'},
    'CSY': {'code': 'cs', 'name': 'Czech: Czech Republic (捷克语:捷克)'},
    'DAN': {'code': 'da', 'name': 'Danish: Denmark (丹麦语:丹麦)'},
    'GER': {'code': 'de-DE', 'name': 'German: Germany (德语:德国)'},
    'DES': {'code': 'de-CH', 'name': 'German: Switzerland (德语:瑞士)'},
    'DEA': {'code': 'de-AT', 'name': 'German: Austria (德语:奥地利)'},
    'DEL': {'code': 'de-LU', 'name': 'German: Luxembourg (德语:卢森堡)'},
    'DEC': {'code': 'de-LI', 'name': 'German: Liechtenstein (德语:列支敦士登)'},
    'ELL': {'code': 'el', 'name': 'Greek: Greece (希腊语:希腊)'},
    'USA': {'code': 'en-US', 'name': 'English: United States (英语:美国)'},
    'ENG': {'code': 'en-GB', 'name': 'English: United Kingdom (英语:英国)'},
    'ENA': {'code': 'en-AU', 'name': 'English: Australia (英语:澳大利亚)'},
    'ENC': {'code': 'en-CA', 'name': 'English: Canada (英语:加拿大)'},
    'ENZ': {'code': 'en-NZ', 'name': 'English: New Zealand (英语:新西兰)'},
    'ENI': {'code': 'en-IE', 'name': 'English: Ireland (英语:爱尔兰)'},
    'ENS': {'code': 'en-ZA', 'name': 'English: South Africa (英语:南非)'},
    'ESP': {'code': 'es-ES', 'name': 'Spanish: Spain (西班牙语:西班牙)'},
    'ESM': {'code': 'es-MX', 'name': 'Spanish: Mexico (西班牙语:墨西哥)'},
    'ESG': {'code': 'es-GT', 'name': 'Spanish: Guatemala (西班牙语:危地马拉)'},
    'ESC': {'code': 'es-CR', 'name': 'Spanish: Costa Rica (西班牙语:哥斯达黎加)'},
    'ESA': {'code': 'es-PA', 'name': 'Spanish: Panama (西班牙语:巴拿马)'},
    'ESD': {'code': 'es-DO', 'name': 'Spanish: Dominican Republic (西班牙语:多米尼加)'},
    'ESV': {'code': 'es-VE', 'name': 'Spanish: Venezuela (西班牙语:委内瑞拉)'},
    'ESO': {'code': 'es-CO', 'name': 'Spanish: Colombia (西班牙语:哥伦比亚)'},
    'ESR': {'code': 'es-PE', 'name': 'Spanish: Peru (西班牙语:秘鲁)'},
    'ESS': {'code': 'es-AR', 'name': 'Spanish: Argentina (西班牙语:阿根廷)'},
    'ESF': {'code': 'es-EC', 'name': 'Spanish: Ecuador (西班牙语:厄瓜多尔)'},
    'ESL': {'code': 'es-CL', 'name': 'Spanish: Chile (西班牙语:智利)'},
    'ESY': {'code': 'es-UY', 'name': 'Spanish: Uruguay (西班牙语:乌拉圭)'},
    'ESZ': {'code': 'es-PY', 'name': 'Spanish: Paraguay (西班牙语:巴拉圭)'},
    'ESB': {'code': 'es-BO', 'name': 'Spanish: Bolivia (西班牙语:玻利维亚)'},
    'ESE': {'code': 'es-SV', 'name': 'Spanish: El Salvador (西班牙语:萨尔瓦多)'},
    'ESH': {'code': 'es-HN', 'name': 'Spanish: Honduras (西班牙语:洪都拉斯)'},
    'ESI': {'code': 'es-NI', 'name': 'Spanish: Nicaragua (西班牙语:尼加拉瓜)'},
    'ESU': {'code': 'es-PR', 'name': 'Spanish: Puerto Rico (西班牙语:波多黎各)'},
    'FIN': {'code': 'fi', 'name': 'Finnish: Finland (芬兰语:芬兰)'},
    'FRA': {'code': 'fr-FR', 'name': 'French: France (法语:法国)'},
    'FRB': {'code': 'fr-BE', 'name': 'French: Belgium (法语:比利时)'},
    'FRC': {'code': 'fr-CA', 'name': 'French: Canada (法语:加拿大)'},
    'FRS': {'code': 'fr-CH', 'name': 'French: Switzerland (法语:瑞士)'},
    'FRL': {'code': 'fr-LU', 'name': 'French: Luxembourg (法语:卢森堡)'},
    'FRM': {'code': 'fr-MC', 'name': 'French: Monaco (法语:摩纳哥)'},
    'HEB': {'code': 'he', 'name': 'Hebrew: Israel (希伯来语:以色列)'},
    'HUN': {'code': 'hu', 'name': 'Hungarian: Hungary (匈牙利语:匈牙利)'},
    'ISL': {'code': 'is', 'name': 'Icelandic: Iceland (冰岛语:冰岛)'},
    'ITA': {'code': 'it-IT', 'name': 'Italian: Italy (意大利语:意大利)'},
    'ITS': {'code': 'it-CH', 'name': 'Italian: Switzerland (意大利语:瑞士)'},
    'JPN': {'code': 'ja', 'name': 'Japanese: Japan (日语:日本)'},
    'KOR': {'code': 'ko', 'name': 'Korean: Korea (韩语:韩国)'},
    'NLD': {'code': 'nl-NL', 'name': 'Dutch: Netherlands (荷兰语:荷兰)'},
    'NLB': {'code': 'nl-BE', 'name': 'Dutch: Belgium (荷兰语:比利时)'},
    'NOR': {'code': 'nb', 'name': 'Norwegian: Norway (Bokmål) (挪威语:挪威)'},
    'NON': {'code': 'nn', 'name': 'Norwegian: Norway (Nynorsk) (挪威语:挪威)'},
    'PLK': {'code': 'pl', 'name': 'Polish: Poland (波兰语:波兰)'},
    'PTB': {'code': 'pt-BR', 'name': 'Portuguese: Brazil (葡萄牙语:巴西)'},
    'PTG': {'code': 'pt-PT', 'name': 'Portuguese: Portugal (葡萄牙语:葡萄牙)'},
    'ROM': {'code': 'ro', 'name': 'Romanian: Romania (罗马尼亚语:罗马尼亚)'},
    'RUS': {'code': 'ru', 'name': 'Russian: Russia (俄语:俄罗斯)'},
    'HRV': {'code': 'hr', 'name': 'Croatian: Croatia (克罗地亚语:克罗地亚)'},
    'SRL': {'code': 'sr-Latn', 'name': 'Serbian: Serbia (Latin) (塞尔维亚语:塞尔维亚)'},
    'SRB': {'code': 'sr-Cyrl', 'name': 'Serbian: Serbia (Cyrillic) (塞尔维亚语:塞尔维亚)'},
    'SKY': {'code': 'sk', 'name': 'Slovak: Slovakia (斯洛伐克语:斯洛伐克)'},
    'SQI': {'code': 'sq', 'name': 'Albanian: Albania (阿尔巴尼亚语:阿尔巴尼亚)'},
    'SVE': {'code': 'sv-SE', 'name': 'Swedish: Sweden (瑞典语:瑞典)'},
    'SVF': {'code': 'sv-FI', 'name': 'Swedish: Finland (瑞典语:芬兰)'},
    'THA': {'code': 'th', 'name': 'Thai: Thailand (泰语:泰国)'},
    'TRK': {'code': 'tr', 'name': 'Turkish: Turkey (土耳其语:土耳其)'},
    'URP': {'code': 'ur-PK', 'name': 'Urdu: Pakistan (乌尔都语:巴基斯坦)'},
    'IND': {'code': 'id', 'name': 'Indonesian: Indonesia (印尼语:印尼)'},
    'UKR': {'code': 'uk', 'name': 'Ukrainian: Ukraine (乌克兰语:乌克兰)'},
    'BEL': {'code': 'be', 'name': 'Belarusian: Belarus (白俄罗斯语:白俄罗斯)'},
    'SLV': {'code': 'sl', 'name': 'Slovene: Slovenia (斯洛文尼亚语:斯洛文尼亚)'},
    'ETI': {'code': 'et', 'name': 'Estonian: Estonia (爱沙尼亚语:爱沙尼亚)'},
    'LVI': {'code': 'lv', 'name': 'Latvian: Latvia (拉脱维亚语:拉脱维亚)'},
    'LTH': {'code': 'lt', 'name': 'Lithuanian: Lithuania (立陶宛语:立陶宛)'},
    'LTC': {'code': 'lt', 'name': 'Classic Lithuanian: Lithuania (立陶宛语:立陶宛)'},
    'FAR': {'code': 'fa', 'name': 'Farsi: Iran (波斯语:伊朗)'},
    'VIT': {'code': 'vi', 'name': 'Vietnamese: Vietnam (越南语:越南)'},
    'HYE': {'code': 'hy', 'name': 'Armenian: Armenia (亚美尼亚语:亚美尼亚)'},
    'AZE': {'code': 'az-Latn', 'name': 'Azeri: Azerbaijan (Latin) (阿塞拜疆语:阿塞拜疆)'},
    'EUQ': {'code': 'eu', 'name': 'Basque: Spain (巴斯克语:西班牙)'},
    'MKI': {'code': 'mk', 'name': 'FYRO Macedonian (马其顿语:马其顿)'},
    'AFK': {'code': 'af', 'name': 'Afrikaans: South Africa (南非语:南非)'},
    'KAT': {'code': 'ka', 'name': 'Georgian: Georgia (格鲁吉亚语:格鲁吉亚)'},
    'FOS': {'code': 'fo', 'name': 'Faeroese: Faeroe Islands (法罗语:法罗群岛)'},
    'HIN': {'code': 'hi', 'name': 'Hindi: India (印地语:印度)'},
    'MSL': {'code': 'ms-MY', 'name': 'Malay: Malaysia (马来语:马来西亚)'},
    'MSB': {'code': 'ms-BN', 'name': 'Malay: Brunei Darussalam (马来语:文莱)'},
    'KAZ': {'code': 'kk', 'name': 'Kazak: Kazakhstan (哈萨克语:哈萨克斯坦)'},
    'SWK': {'code': 'sw', 'name': 'Swahili: Kenya (斯瓦希里语:肯尼亚)'},
    'UZB': {'code': 'uz-Latn', 'name': 'Uzbek: Uzbekistan (Latin) (乌兹别克语:乌兹别克斯坦)'},
    'TAT': {'code': 'tt', 'name': 'Tatar: Tatarstan (鞑靼语:鞑靼斯坦)'},
    'BEN': {'code': 'bn', 'name': 'Bengali: India (孟加拉语:印度)'},
    'PAN': {'code': 'pa', 'name': 'Punjabi: India (旁遮普语:印度)'},
    'GUJ': {'code': 'gu', 'name': 'Gujarati: India (古吉拉特语:印度)'},
    'ORI': {'code': 'or', 'name': 'Oriya: India (奥里亚语:印度)'},
    'TAM': {'code': 'ta', 'name': 'Tamil: India (泰米尔语:印度)'},
    'TEL': {'code': 'te', 'name': 'Telugu: India (泰卢固语:印度)'},
    'KAN': {'code': 'kn', 'name': 'Kannada: India (卡纳达语:印度)'},
    'MAL': {'code': 'ml', 'name': 'Malayalam: India (马拉雅拉姆语:印度)'},
    'ASM': {'code': 'as', 'name': 'Assamese: India (阿萨姆语:印度)'},
    'MAR': {'code': 'mr', 'name': 'Marathi: India (马拉地语:印度)'},
    'SAN': {'code': 'sa', 'name': 'Sanskrit: India (梵语:印度)'},
    'KOK': {'code': 'kok', 'name': 'Konkani: India (孔卡尼语:印度)'}
}

# RES输出时追加的语言标识: 语言代码 → 本机语言名+中文国家名
LANG_NATIVE_NAME = {
    'CHS': '中文(简体) 中国', 'CHT': '中文(繁體) 台湾', 'ZHH': '中文(香港) 香港',
    'ZHI': '中文(台湾) 台湾', 'ZHM': '中文(澳门) 澳门',
    'ENG': 'English 美国', 'USA': 'English 美国', 'ENA': 'English 美国', 'ENC': 'English 英国',
    'ENZ': 'English 新西兰', 'ENI': 'English 印度', 'ENS': 'English 新加坡',
    'GER': 'Deutsch 德国', 'DES': 'Deutsch 瑞士', 'DEA': 'Deutsch 奥地利', 'DEL': 'Deutsch 列支敦士登', 'DEC': 'Deutsch 卢森堡',
    'JPN': '日本語 日本', 'KOR': '한국어 韩国', 'FRA': 'Français 法国', 'FRB': 'Français 比利时',
    'FRC': 'Français 加拿大', 'FRS': 'Français 瑞士', 'FRL': 'Français 卢森堡', 'FRM': 'Français 摩纳哥',
    'ITA': 'Italiano 意大利', 'ITS': 'Italiano 瑞士',
    'ESP': 'Español 西班牙', 'ESM': 'Español 墨西哥', 'ESG': 'Español 危地马拉', 'ESC': 'Español 哥斯达黎加',
    'ESA': 'Español 巴拿马', 'ESD': 'Español 多米尼加共和国', 'ESV': 'Español 委内瑞拉', 'ESO': 'Español 哥伦比亚',
    'ESR': 'Español 秘鲁', 'ESS': 'Español 阿根廷', 'ESF': 'Español 厄瓜多尔', 'ESL': 'Español 智利',
    'ESY': 'Español 乌拉圭', 'ESB': 'Español 玻利维亚', 'ESE': 'Español 巴拉圭', 'ESH': 'Esp萨尔瓦多',
    'ESN': 'Español 尼加拉瓜', 'ESU': 'Español 波多黎各', 'ESP2': 'Español 美国',
    'RUS': 'Русский 俄罗斯', 'PTG': 'Português 葡萄牙', 'PTB': 'Português 巴西',
    'TRK': 'Türkçe 土耳其', 'PLK': 'Polski 波兰', 'VIT': 'Tiếng Việt 越南',
    'CAT': 'Català 西班牙', 'THA': 'ไทย 泰国', 'DAN': 'Dansk 丹麦', 'NON': 'Norsk 挪威',
    'SVE': 'Svenska 瑞典', 'FIN': 'Suomi 芬兰', 'NLN': 'Nederlands 荷兰', 'NLB': 'Nederlands 比利时',
    'CSY': 'Čeština 捷克', 'SKY': 'Slovenčina 斯洛伐克', 'HUN': 'Magyar 匈牙利', 'ELL': 'Ελληνικά 希腊',
    'ARA': 'العربية 沙特阿拉伯', 'ARL': 'العربية 利比亚', 'ARG': 'العربية 阿尔及利亚', 'ARM': 'العربية 摩洛哥',
    'ART': 'الع العربية 伊拉克', 'ARO': 'العربية 阿曼', 'ARY': 'العربية 埃及', 'ARS': 'العربية 叙利亚',
    'ARJ': 'العربية 约旦', 'ARB': 'العربية 黎巴嫩', 'ARK': 'العربية 科威特', 'ARU': 'العربية 阿联酋',
    'ARH': 'العربية 巴林', 'ARQ': 'العربية 卡塔尔',
    'BGR': 'Български 保加利亚', 'UKR': 'Українська 乌克兰', 'BEL': 'Беларуская 白俄罗斯',
    'KAZ': 'Қазақша 哈萨克斯坦', 'SRB': 'Српски 塞尔维亚', 'HRV': 'Hrvatski 克罗地亚', 'SLV': 'Slovenščina 斯洛文尼亚',
    'EST': 'Eesti 爱沙尼亚', 'LVI': 'Latviešu 拉脱维亚', 'LTH': 'Lietuvių 立陶宛', 'LTC': 'Lietuvių 立陶宛',
    'ROM': 'Română 罗马尼亚', 'ROM2': 'Română 摩尔多瓦', 'IND': 'Bahasa Indonesia 印尼',
    'HEB': 'עברית 以色列', 'FAR': 'فارسی 伊朗', 'URD': 'اردو 巴基斯坦',
    'HYE': 'Հայերեն 亚美尼亚', 'KAT': 'ქართული 格鲁吉亚', 'AZE': 'Azərbaycan 阿塞拜疆',
    'HIN': 'हिन्दी 印度', 'BEN': 'বাংলা 孟加拉', 'PAN': 'ਪੰਜਾਬੀ 印度', 'GUJ': 'ગુજરાતી 印度',
    'ORI': 'ଓଡ଼ିଆ 印度', 'TAM': 'தமிழ் 印度', 'TEL': 'తెలుగు 印度', 'KAN': 'ಕನ್ನಡ 印度',
    'MAL': 'മലയാളം 印度', 'ASM': 'অসমীয়া 印度', 'MAR': 'मराठी 印度', 'SAN': 'संस्कृतम् 印度',
    'KOK': 'कोंकणी 印度', 'MSL': 'Bahasa Melayu 马来西亚', 'MSB': 'Bahasa Melayu 文莱',
    'SWK': 'Kiswahili 坦桑尼亚', 'UZB': 'O\'zbek 乌兹别克斯坦', 'TAT': 'Татарча 俄罗斯',
    'MKI': 'Македонски 北马其顿', 'AFK': 'Afrikaans 南非', 'FOS': 'Føroyskt 法罗群岛',
}

# 默认显示的语言列表（11种常用语言）
DEFAULT_LANGS = ['CAT', 'CHT', 'ESP', 'GER', 'ITA', 'KOR', 'PLK', 'PTG', 'RUS', 'TRK', 'VIT']

# 品牌名称集合 - 这些名称不翻译
BRAND_NAMES = {'禾川', '高创', '松下', '汇川', '迪维迅', '台达', '雷赛', '信捷', 'H系列'}

# 配置文件名
CONFIG_FILE = 'translation_config.json'


class TranslationApp:
    """
    自动化翻译处理工具主类
    提供可视化用户界面，支持XML文件翻译和多语言输出
    """
    
    def __init__(self, root):
        """
        初始化方法
        :param root: Tkinter主窗口对象
        """
        self.root = root
        self.root.title("自动化翻译处理工具")
        self.root.geometry("1000x800")
        
        # 初始化实例变量
        self.source_files = []           # 源文件列表
        self.source_folder = ""         # 源文件夹路径（CLI兼容）
        self.source_folders = []        # 已添加的文件夹路径列表（统一UI）
        self.selected_langs = []        # 选中的语言列表
        self.translation_table = {}     # 翻译表字典
        self.output_dir = os.path.dirname(os.path.abspath(__file__))  # 输出目录
        self.table_dir = os.path.dirname(os.path.abspath(__file__))   # 翻译表保存目录
        self.default_langs = DEFAULT_LANGS.copy()  # 默认语言列表
        self.show_all_langs = False     # 是否显示所有语言
        self.translation_complete = False  # 翻译是否完成
        self.headless = False  # 是否为无界面模式（命令行）
        self.logger = None
        
        # DiskC工作流模式
        self.diskc_mode = False       # 是否为DiskC工作流模式
        self.diskc_root = ""          # DiskC根目录路径
        self.diskc_res_source = ""    # CHS.res解压后的prepared路径
        self.diskc_xml_map = {}       # prepared_path -> 相对于String目录的子路径
        
        # API配置变量
        self.api_type = 'google'  # 当前使用的翻译API类型：google 或 baidu
        self.baidu_app_id = ''   # 百度翻译API App ID
        self.baidu_secret_key = ''  # 百度翻译API Secret Key
        
        # 加载配置、创建界面、加载翻译表
        self.load_config()
        self.create_widgets()
        self.load_translation_table()
    
    def load_config(self):
        """
        加载配置文件
        从translation_config.json读取用户自定义的默认语言设置和API配置
        """
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    if 'default_langs' in config:
                        self.default_langs = config['default_langs']
                    if 'api_type' in config:
                        self.api_type = config['api_type']
                    if 'baidu_app_id' in config:
                        self.baidu_app_id = config['baidu_app_id']
                    if 'baidu_secret_key' in config:
                        self.baidu_secret_key = config['baidu_secret_key']
            except Exception as e:
                self.log(f"加载配置失败: {e}")
                self._ensure_logger()
                if self.logger:
                    self.logger.exception(f"加载配置失败: {e}")

    def _ensure_logger(self, log_file=None):
        """
        初始化文件日志（可选）
        """
        if self.logger:
            return
        try:
            self.logger = logging.getLogger('translation_app')
            self.logger.setLevel(logging.DEBUG)
            fmt = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
            # 控制台处理器
            ch = logging.StreamHandler()
            ch.setFormatter(fmt)
            self.logger.addHandler(ch)
            # 可选文件处理器
            if log_file:
                fh = logging.FileHandler(log_file, encoding='utf-8')
                fh.setFormatter(fmt)
                self.logger.addHandler(fh)
        except Exception:
            self.logger = None
    
    def save_config(self):
        """
        保存配置文件
        将当前默认语言设置和API配置保存到translation_config.json
        """
        config = {
            'default_langs': self.default_langs,
            'api_type': self.api_type,
            'baidu_app_id': self.baidu_app_id,
            'baidu_secret_key': self.baidu_secret_key
        }
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
    
    def create_widgets(self):
        self.root.title("Translate Pro")
        self.root.geometry("1280x850")
        self.root.configure(bg='#F5F6FA')
        
        style = ttk.Style()
        style.theme_use('clam')
        
        style.configure('TFrame', background='#F5F6FA')
        style.configure('Card.TFrame', background='white')
        style.configure('Sidebar.TFrame', background='#2D2B4E')
        style.configure('Stats.TFrame', background='white')
        style.configure('Header.TLabel', background='#2D2B4E', foreground='white', font=('Segoe UI', 11, 'bold'))
        style.configure('Nav.TLabel', background='#2D2B4E', foreground='#A5A0C0', font=('Segoe UI', 10))
        style.configure('NavActive.TLabel', background='#2D2B4E', foreground='white', font=('Segoe UI', 10, 'bold'))
        style.configure('Title.TLabel', background='#F5F6FA', foreground='#2D2B4E', font=('Segoe UI', 13, 'bold'))
        style.configure('Subtitle.TLabel', background='#F5F6FA', foreground='#6B6B8D', font=('Segoe UI', 9))
        style.configure('StatValue.TLabel', background='white', foreground='#2D2B4E', font=('Segoe UI', 24, 'bold'))
        style.configure('StatLabel.TLabel', background='white', foreground='#8E8EA0', font=('Segoe UI', 9))
        style.configure('Accent.TButton', font=('Segoe UI', 10, 'bold'))
        style.configure('TButton', font=('Segoe UI', 9), padding=6)
        style.configure('TCheckbutton', background='#F5F6FA', font=('Segoe UI', 9))
        style.configure('TProgressbar', thickness=8, troughcolor='#E8E8F0', background='#6C63FF')
        style.configure('Card.TLabelframe', background='white')
        style.configure('Card.TLabelframe.Label', background='white', foreground='#2D2B4E', font=('Segoe UI', 11, 'bold'))
        
        self.root.grid_columnconfigure(1, weight=1)
        self.root.grid_rowconfigure(0, weight=1)
        
        sidebar = ttk.Frame(self.root, style='Sidebar.TFrame', width=200)
        sidebar.grid(row=0, column=0, sticky='ns')
        sidebar.grid_propagate(False)
        
        ttk.Label(sidebar, text="Translate", style='Header.TLabel').pack(pady=(30, 8), anchor='w', padx=24)
        ttk.Label(sidebar, text="Pro", style='Header.TLabel', foreground='#6C63FF').pack(anchor='w', padx=24)
        
        ttk.Separator(sidebar, orient='horizontal').pack(fill='x', pady=20, padx=16)
        
        nav_items = [
            ("📁", "文件管理", "files"),
            ("🌐", "语言选择", "langs"),
            ("📊", "处理进度", "progress"),
            ("📝", "处理日志", "logs"),
            ("⚙️", "设置选项", "settings")
        ]
        
        self.nav_buttons = {}
        for icon, text, name in nav_items:
            btn_frame = ttk.Frame(sidebar, style='Sidebar.TFrame')
            btn_frame.pack(fill='x', padx=12, pady=2)
            
            lbl = ttk.Label(btn_frame, text=f"{icon}  {text}", style='Nav.TLabel',
                           cursor='hand2')
            lbl.pack(fill='x', pady=8, padx=12)
            lbl.bind('<Enter>', lambda e, l=lbl: l.configure(style='NavActive.TLabel'))
            lbl.bind('<Leave>', lambda e, l=lbl: l.configure(style='Nav.TLabel'))
            self.nav_buttons[name] = btn_frame
        
        ttk.Frame(sidebar, style='Sidebar.TFrame').pack(side='bottom', fill='x', pady=20)
        
        main_content = ttk.Frame(self.root, style='TFrame')
        main_content.grid(row=0, column=1, sticky='nsew', padx=16, pady=16)
        main_content.grid_columnconfigure(0, weight=1)
        main_content.grid_rowconfigure(3, weight=1)
        
        header_frame = ttk.Frame(main_content, style='TFrame')
        header_frame.grid(row=0, column=0, sticky='ew', pady=(0, 16))
        header_frame.grid_columnconfigure(1, weight=1)
        
        ttk.Label(header_frame, text="翻译工作台", style='Title.TLabel').grid(row=0, column=0, sticky='w')
        ttk.Label(header_frame, text="支持 XML/RES 文件 · 100+ 语言 · 智能缩写", style='Subtitle.TLabel').grid(row=1, column=0, sticky='w', pady=(2, 0))
        
        action_frame = ttk.Frame(header_frame, style='TFrame')
        action_frame.grid(row=0, column=2, rowspan=2, sticky='e')
        
        api_select = ttk.Frame(action_frame, style='TFrame')
        api_select.pack(side='left', padx=(0, 8))
        
        ttk.Label(api_select, text="API:", font=('Segoe UI', 9)).pack(side='left')
        self.api_type_var = tk.StringVar(value=self.api_type)
        api_combo = ttk.Combobox(api_select, textvariable=self.api_type_var, values=['google', 'baidu'],
                                 state='readonly', width=7, font=('Segoe UI', 9))
        api_combo.pack(side='left', padx=4)
        api_combo.bind('<<ComboboxSelected>>', lambda e: self.on_api_type_changed())
        ttk.Button(api_select, text="配置", command=self.configure_api).pack(side='left')
        
        ttk.Separator(action_frame, orient='vertical').pack(side='left', fill='y', padx=6)
        
        self.start_btn = ttk.Button(action_frame, text="▶ 开始翻译", command=self.start_translation, style='Accent.TButton')
        self.start_btn.pack(side='left', padx=4)
        
        self.confirm_btn = ttk.Button(action_frame, text="✓ 确认生成", command=self.confirm_and_generate, state='disabled')
        self.confirm_btn.pack(side='left', padx=4)
        
        content_area = ttk.Frame(main_content, style='TFrame')
        content_area.grid(row=1, column=0, sticky='nsew')
        content_area.grid_columnconfigure(0, weight=1)
        content_area.grid_columnconfigure(1, weight=1)
        content_area.grid_rowconfigure(1, weight=1)
        
        files_card = ttk.LabelFrame(content_area, text=" 📂 输入源文件 ", style='Card.TLabelframe', padding=12)
        files_card.grid(row=0, column=0, columnspan=2, sticky='ew', pady=(0, 12))
        files_card.grid_columnconfigure(1, weight=1)
        
        toolbar = ttk.Frame(files_card, style='Card.TFrame')
        toolbar.grid(row=0, column=0, columnspan=2, sticky='ew', pady=(0, 8))
        
        ttk.Button(toolbar, text="+ 添加文件夹", command=self.add_folder).pack(side='left', padx=2)
        ttk.Button(toolbar, text="+ 添加文件", command=self.add_files).pack(side='left', padx=2)
        ttk.Button(toolbar, text="DiskC 工作流", command=self.setup_diskc_sources).pack(side='left', padx=2)
        ttk.Button(toolbar, text="移除选中", command=self.remove_files).pack(side='left', padx=2)
        ttk.Button(toolbar, text="清空列表", command=self.clear_files).pack(side='left', padx=2)
        
        ttk.Separator(toolbar, orient='vertical').pack(side='left', fill='y', padx=8, pady=2)
        
        self.pack_var = tk.BooleanVar()
        ttk.Checkbutton(toolbar, text="打包 .res", variable=self.pack_var).pack(side='left', padx=4)
        
        ttk.Button(toolbar, text="翻译表", command=self.view_translation_table).pack(side='left', padx=2)
        ttk.Button(toolbar, text="导出Excel", command=self.export_to_excel).pack(side='left', padx=2)
        ttk.Button(toolbar, text="导入Excel", command=self.import_from_excel).pack(side='left', padx=2)
        ttk.Button(toolbar, text="导出日志", command=self.export_log).pack(side='left', padx=2)
        
        list_container = ttk.Frame(files_card, style='Card.TFrame')
        list_container.grid(row=1, column=0, columnspan=2, sticky='nsew')
        files_card.grid_rowconfigure(1, weight=1)
        
        self.file_listbox = tk.Listbox(list_container, height=4, bg='#FAFBFE', fg='#2D2B4E',
                                       font=('Segoe UI', 9), selectbackground='#6C63FF', selectforeground='white',
                                       borderwidth=0, highlightthickness=1, highlightbackground='#E8E8F0')
        self.file_listbox.pack(side='left', fill='both', expand=True)
        
        file_scrollbar = ttk.Scrollbar(list_container, orient='vertical', command=self.file_listbox.yview)
        file_scrollbar.pack(side='right', fill='y')
        self.file_listbox.configure(yscrollcommand=file_scrollbar.set)
        
        lang_card = ttk.LabelFrame(content_area, text=" 🌍 目标语言 ", style='Card.TLabelframe', padding=12)
        lang_card.grid(row=1, column=0, sticky='nsew', padx=(0, 6))
        
        lang_toolbar = ttk.Frame(lang_card, style='Card.TFrame')
        lang_toolbar.grid(row=0, column=0, sticky='ew', pady=(0, 8))
        
        self.lang_show_var = tk.BooleanVar(value=self.show_all_langs)
        ttk.Checkbutton(lang_toolbar, text="显示全部语言", variable=self.lang_show_var,
                       command=self.toggle_lang_display).pack(side='left', padx=4)
        
        self.main_lang_search_var = tk.StringVar()
        search_entry = ttk.Entry(lang_toolbar, textvariable=self.main_lang_search_var, width=14,
                                  font=('Segoe UI', 9))
        search_entry.pack(side='left', padx=(8, 4))
        search_entry.insert(0, '🔍 搜索...')
        search_entry.configure(foreground='#A5A0C0')
        
        def _on_focus_in(e):
            if self.main_lang_search_var.get() == '🔍 搜索...':
                search_entry.delete(0, tk.END)
                search_entry.configure(foreground='#2D2B4E')
        
        def _on_focus_out(e):
            if not self.main_lang_search_var.get().strip():
                search_entry.insert(0, '🔍 搜索...')
                search_entry.configure(foreground='#A5A0C0')
        
        search_entry.bind('<FocusIn>', _on_focus_in)
        search_entry.bind('<FocusOut>', _on_focus_out)
        self.main_lang_search_var.trace('w', lambda *args: self._filter_main_langs())
        
        ttk.Button(lang_toolbar, text="设置默认", command=self.set_default_langs).pack(side='left', padx=4)
        
        btn_frame = ttk.Frame(lang_toolbar, style='Card.TFrame')
        btn_frame.pack(side='right')
        ttk.Button(btn_frame, text="全选", command=self.select_all_langs).pack(side='left', padx=2)
        ttk.Button(btn_frame, text="清空", command=self.deselect_all_langs).pack(side='left', padx=2)
        
        lang_canvas_frame = ttk.Frame(lang_card, style='Card.TFrame')
        lang_canvas_frame.grid(row=1, column=0, sticky='nsew')
        lang_card.grid_rowconfigure(1, weight=1)
        
        canvas = tk.Canvas(lang_canvas_frame, bg='white', highlightthickness=0, width=340, height=260)
        scrollbar_y = ttk.Scrollbar(lang_canvas_frame, orient='vertical', command=canvas.yview)
        scrollable_lang = ttk.Frame(canvas, style='Card.TFrame')
        
        scrollable_lang.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox('all')))
        canvas.create_window((0, 0), window=scrollable_lang, anchor='nw')
        canvas.configure(yscrollcommand=scrollbar_y.set)
        
        def on_mousewheel(e):
            canvas.yview_scroll(int(-1 * (e.delta / 120)), 'units')
        canvas.bind_all('<MouseWheel>', on_mousewheel, add='+')
        
        canvas.pack(side='left', fill='both', expand=True)
        scrollbar_y.pack(side='right', fill='y')
        
        self.lang_frame = scrollable_lang
        
        right_panel = ttk.Frame(content_area, style='TFrame')
        right_panel.grid(row=1, column=1, sticky='nsew', padx=(6, 0))
        right_panel.grid_rowconfigure(1, weight=1)
        
        stats_card = ttk.LabelFrame(right_panel, text=" 📊 统计信息 ", style='Card.TLabelframe', padding=16)
        stats_card.grid(row=0, column=0, sticky='ew', pady=(0, 12))
        
        self.stat_file_count = ttk.Label(stats_card, text="--", style='StatValue.TLabel')
        self.stat_file_count.pack(anchor='center')
        ttk.Label(stats_card, text="已添加文件", style='StatLabel.TLabel').pack(anchor='center', pady=(0, 12))
        
        stat_mid = ttk.Frame(stats_card, style='Stats.TFrame')
        stat_mid.pack(fill='x', pady=8)
        
        self.stat_lang_count = ttk.Label(stat_mid, text="--", style='StatValue.TLabel', font=('Segoe UI', 16, 'bold'))
        self.stat_lang_count.pack(side='left')
        ttk.Label(stat_mid, text="\n已选语言", style='StatLabel.TLabel').pack(side='left', padx=(8, 0))
        
        progress_card = ttk.LabelFrame(right_panel, text=" ⏳ 处理进度 ", style='Card.TLabelframe', padding=16)
        progress_card.grid(row=1, column=0, sticky='nsew')
        
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(progress_card, variable=self.progress_var, maximum=100,
                                           mode='determinate', length=200)
        self.progress_bar.pack(fill='x', pady=(0, 8))
        
        pct_frame = ttk.Frame(progress_card, style='Stats.TFrame')
        pct_frame.pack()
        self.progress_label = ttk.Label(pct_frame, text="0%", font=('Segoe UI', 14, 'bold'),
                                        foreground='#6C63FF')
        self.progress_label.pack(side='left')
        ttk.Label(pct_frame, text="  就绪", style='Subtitle.TLabel').pack(side='left')
        
        log_card = ttk.LabelFrame(main_content, text=" 📋 处理日志 ", style='Card.TLabelframe', padding=12)
        log_card.grid(row=2, column=0, sticky='nsew', pady=(12, 0))
        main_content.grid_rowconfigure(2, weight=1)
        
        self.log_text = scrolledtext.ScrolledText(log_card, height=10, wrap=tk.WORD,
                                                   font=('Consolas', 9), bg='#FAFBFE', fg='#4A4A68',
                                                   borderwidth=0, highlightthickness=0)
        self.log_text.pack(fill='both', expand=True)
        self.log_text.insert(tk.END, "欢迎使用 Translate Pro！\n")
        self.log_text.insert(tk.END, "请添加输入源文件并选择目标语言，然后点击「开始翻译」。\n")
        
        stats_panel = ttk.Frame(self.root, style='Stats.TFrame', width=220)
        stats_panel.grid(row=0, column=2, sticky='ns', padx=(0, 16), pady=16)
        stats_panel.grid_propagate(False)
        
        outer_stats = ttk.Frame(stats_panel, style='Stats.TFrame', padding=20)
        outer_stats.pack(fill='both', expand=True)
        
        ttk.Label(outer_stats, text="Statistic", font=('Segoe UI', 14, 'bold'),
                  foreground='#2D2B4E', background='white').pack(anchor='w', pady=(0, 20))
        
        stat_items = [
            ("downloads", "本周处理", "0", "次"),
            ("space", "可用空间", "--", ""),
            ("shared", "翻译条目", "0", "条")
        ]
        
        self.stats_refs = {}
        for icon_id, label, default_val, unit in stat_items:
            item_frame = ttk.Frame(outer_stats, style='Stats.TFrame')
            item_frame.pack(fill='x', pady=10)
            
            left_part = ttk.Frame(item_frame, style='Stats.TFrame')
            left_part.pack(side='left')
            
            val_lbl = ttk.Label(left_part, text=default_val, style='StatValue.TLabel',
                                font=('Segoe UI', 18, 'bold'))
            val_lbl.pack(anchor='w')
            ttk.Label(left_part, text=label, style='StatLabel.TLabel').pack(anchor='w')
            
            self.stats_refs[icon_id] = val_lbl
            
            if unit:
                ttk.Label(item_frame, text=f"\n{unit}", style='StatLabel.TLabel').pack(side='left', padx=(8, 0))
        
        ttk.Separator(outer_stats, orient='horizontal').pack(fill='x', pady=16)
        
        tip_frame = ttk.Frame(outer_stats, style='Stats.TFrame')
        tip_frame.pack(fill='x')
        ttk.Label(tip_frame, text="💡 提示", font=('Segoe UI', 10, 'bold'),
                  foreground='#2D2B4E', background='white').pack(anchor='w')
        ttk.Label(tip_frame, text="智能缩写将自动优化\n超长翻译文本的长度",
                  font=('Segoe UI', 9), foreground='#8E8EA0', background='white',
                  justify='left').pack(anchor='w', pady=(4, 0))
        
        self.update_lang_display()
    
    def add_folder(self):
        folder = filedialog.askdirectory(title="选择输入文件夹")
        if not folder:
            return
        self.source_folders.append(folder)
        if not self.source_files:
            self.output_dir = os.path.dirname(folder)
            self.log(f"输出目录已设置为: {self.output_dir}")
        count = 0
        for root_dir, dirs, files in os.walk(folder):
            for file in files:
                file_path = os.path.join(root_dir, file)
                prepared = self.prepare_file(file_path)
                if prepared and prepared not in self.source_files:
                    self.source_files.append(prepared)
                    tag = self._file_tag(prepared)
                    self.file_listbox.insert(tk.END, f"{tag} {prepared}")
                    count += 1
        self.log(f"从文件夹添加 {count} 个文件: {folder}")

    def refresh_file_list(self):
        self.source_files = []
        self.file_listbox.delete(0, tk.END)
        all_folders = list(self.source_folders)
        if self.source_folder and self.source_folder not in all_folders:
            all_folders.append(self.source_folder)
        if not all_folders:
            self.log("无已添加文件夹，请先添加文件夹或文件")
            return
        for folder in all_folders:
            if os.path.isdir(folder):
                count = 0
                for root_dir, dirs, files in os.walk(folder):
                    for file in files:
                        file_path = os.path.join(root_dir, file)
                        prepared = self.prepare_file(file_path)
                        if prepared and prepared not in self.source_files:
                            self.source_files.append(prepared)
                            tag = self._file_tag(prepared)
                            self.file_listbox.insert(tk.END, f"{tag} {prepared}")
                            count += 1
                self.log(f"刷新文件夹: {os.path.basename(folder)} → {count} 个文件")
        self.log(f"总计扫描到 {len(self.source_files)} 个可处理文件")

    def setup_diskc_sources(self, diskc_root=None):
        if diskc_root is None:
            folder = filedialog.askdirectory(title="选择DiskC根目录")
            if not folder:
                return
            diskc_root = folder

        self.diskc_root = diskc_root
        self.diskc_mode = True
        self.diskc_res_source = ""
        self.diskc_xml_map = {}
        self.source_files = []
        self.file_listbox.delete(0, tk.END)
        self.source_folders = []
        self.source_folder = diskc_root
        self.output_dir = diskc_root

        res_path = os.path.join(diskc_root, 'OpenCNC', 'Bin', 'Language', 'CHS.res')
        prepared_res = self.prepare_file(res_path)
        if prepared_res:
            self.diskc_res_source = prepared_res
            self.source_files.append(prepared_res)
            self.file_listbox.insert(tk.END, f"[RES] {os.path.basename(res_path)} -> {prepared_res}")
            self.log(f"DiskC工作流: 加载 {res_path}")
        else:
            self.log(f"DiskC工作流: 未找到 {res_path}")

        string_dir = os.path.join(diskc_root, 'OpenCnc Shared', 'OCRes', 'CHS', 'String')
        if os.path.isdir(string_dir):
            xml_count = 0
            for root_dir, dirs, files in os.walk(string_dir):
                for file in files:
                    if file.lower().endswith('.xml'):
                        full_path = os.path.join(root_dir, file)
                        prepared = self.prepare_file(full_path)
                        if prepared:
                            rel_path = os.path.relpath(prepared, string_dir)
                            self.diskc_xml_map[prepared] = rel_path
                            self.source_files.append(prepared)
                            self.file_listbox.insert(tk.END, f"[XML] String/{rel_path}")
                            xml_count += 1
            self.log(f"DiskC工作流: 从String目录加载 {xml_count} 个XML文件")
        else:
            self.log(f"DiskC工作流: 未找到目录 {string_dir}")

        msg = f"DiskC工作流已就绪: {len(self.source_files)} 个文件"
        if not self.headless:
            messagebox.showinfo("DiskC工作流", msg)
        self.log(msg)

    def prepare_file(self, file_path):
        """
        准备单个文件：自动识别类型，RES文件解压为无后缀名的XML文件并返回新路径
        返回可供后续处理的文件路径（XML文本文件路径），失败返回None
        """
        try:
            if file_path.lower().endswith('.res'):
                # 将RES解压写入到输出目录下的临时文件（无扩展名）
                base = os.path.splitext(os.path.basename(file_path))[0]
                conv_dir = os.path.join(self.output_dir, '_res_converted')
                os.makedirs(conv_dir, exist_ok=True)
                target_path = os.path.join(conv_dir, base)  # 无后缀名
                if os.path.exists(target_path) and os.path.getmtime(target_path) >= os.path.getmtime(file_path):
                    # 验证缓存文件是否为有效XML内容
                    try:
                        with open(target_path, 'r', encoding='utf-8') as tf:
                            preview = tf.read(200).strip()
                        if preview and (preview.startswith('<?xml') or preview.startswith('<')):
                            return target_path
                        else:
                            self.log(f"缓存文件 {os.path.basename(target_path)} 内容无效，重新解压")
                    except Exception:
                        self.log(f"缓存文件 {os.path.basename(target_path)} 读取失败，重新解压")
                content = self.decompress_res(file_path)
                if content is None:
                    self.log(f"RES解压失败: {file_path}")
                    if self.logger:
                        self.logger.error(f"RES解压失败: {file_path}")
                    return None
                with open(target_path, 'w', encoding='utf-8') as f:
                    f.write(content)
                self.log(f"已转换RES -> 无后缀XML: {os.path.basename(target_path)}")
                return target_path
            else:
                # 仅接受XML或无后缀（可能是已转换的RES）
                if file_path.lower().endswith('.xml') or not os.path.splitext(file_path)[1]:
                    return file_path
                # 其它类型不支持
                self.log(f"跳过不支持的文件类型: {file_path}")
                if self.logger:
                    self.logger.warning(f"跳过不支持的文件类型: {file_path}")
                return None
        except Exception as e:
            self.log(f"准备文件失败: {file_path} -> {e}")
            if self.logger:
                self.logger.exception(e)
            return None
    
    def add_files(self):
        files = filedialog.askopenfilenames(title="选择XML或RES文件",
            filetypes=[("支持的源文件", "*.xml;*.res"), ("XML文件", "*.xml"), ("RES文件", "*.res"), ("所有文件", "*.*")])
        if not files:
            return
        if not self.source_files:
            first_file_dir = os.path.dirname(files[0])
            self.output_dir = os.path.dirname(first_file_dir)
            self.log(f"输出目录已设置为: {self.output_dir}")
        count = 0
        for file in files:
            prepared = self.prepare_file(file)
            if prepared and prepared not in self.source_files:
                self.source_files.append(prepared)
                tag = self._file_tag(prepared)
                self.file_listbox.insert(tk.END, f"{tag} {prepared}")
                count += 1
            else:
                self.log(f"跳过文件: {file}")
        if count:
            self.log(f"添加 {count} 个文件")

    @staticmethod
    def _file_tag(file_path):
        lower = file_path.lower()
        if lower.endswith('.res'):
            return '[RES]'
        elif '_res_converted' in lower:
            return '[RES→XML]'
        elif lower.endswith('.xml'):
            return '[XML]'
        elif not os.path.splitext(file_path)[1]:
            return '[无后缀]'
        return '[?]'
    
    def remove_files(self):
        """
        从列表中移除选中的文件
        """
        selected = self.file_listbox.curselection()
        for idx in reversed(selected):
            self.file_listbox.delete(idx)
            del self.source_files[idx]
    
    def clear_files(self):
        self.file_listbox.delete(0, tk.END)
        self.source_files = []
        self.source_folders = []
    
    def toggle_lang_display(self):
        """
        切换语言显示模式
        显示所有语言或仅显示默认语言
        """
        self.show_all_langs = self.lang_show_var.get()
        self.update_lang_display()
    
    def update_lang_display(self):
        """
        更新语言复选框显示
        根据当前模式显示相应的语言选项
        """
        # 清除现有组件
        for widget in self.lang_frame.winfo_children():
            widget.destroy()
        
        self.lang_vars = {}
        self.main_lang_frames = {}
        
        display_langs = LANG_MAP if self.show_all_langs else {k: v for k, v in LANG_MAP.items() if k in self.default_langs}
        
        col = 0
        row_lang = 0
        for lang, info in sorted(display_langs.items()):
            var = tk.BooleanVar()
            chk = ttk.Checkbutton(self.lang_frame, text=f"{lang} - {info['name']}", variable=var)
            chk.grid(row=row_lang, column=col, sticky=tk.W, padx=5, pady=2)
            self.lang_vars[lang] = var
            self.main_lang_frames[lang] = chk
            col += 1
            if col >= 2:
                col = 0
                row_lang += 1
        
        self.lang_frame.update_idletasks()
        self.update_stats()
    
    def update_stats(self):
        if hasattr(self, 'stat_file_count'):
            self.stat_file_count.configure(text=str(len(self.source_files)))
        if hasattr(self, 'stat_lang_count'):
            selected = len([l for l, v in self.lang_vars.items() if v.get()]) if hasattr(self, 'lang_vars') else 0
            self.stat_lang_count.configure(text=str(selected))
        if hasattr(self, 'stats_refs') and 'shared' in self.stats_refs:
            total = len(self.translation_table)
            self.stats_refs['shared'].configure(text=str(total))
    
    def _filter_main_langs(self):
        if not hasattr(self, 'main_lang_frames'):
            return
        txt = self.main_lang_search_var.get().strip()
        if txt == '🔍 搜索...':
            txt = ''
        txt = txt.lower()
        import re as _re
        
        for code, chk in self.main_lang_frames.items():
            info = LANG_MAP.get(code, {})
            name = info.get('name', '')
            cn = ' '.join(_re.findall(r'[\u4e00-\u9fff]+', name)).lower()
            
            if not txt or txt in code.lower() or txt in name.lower() or txt in cn:
                chk.grid()
            else:
                chk.grid_remove()
    
    def select_all_langs(self):
        """
        全选所有语言
        """
        for var in self.lang_vars.values():
            var.set(True)
    
    def deselect_all_langs(self):
        """
        取消全选所有语言
        """
        for var in self.lang_vars.values():
            var.set(False)
    
    def set_default_langs(self):
        top = tk.Toplevel(self.root)
        top.title("设置默认语言")
        top.geometry("680x650")
        top.configure(bg='#F5F6FA')
        top.transient(self.root)
        top.grab_set()
        
        dlg_style = ttk.Style()
        if 'Dialog.TFrame' not in dlg_style.theme_names():
            dlg_style.configure('Dialog.TFrame', background='#F5F6FA')
            dlg_style.configure('DialogCard.TFrame', background='white')
            dlg_style.configure('DialogTitle.TLabel', background='#F5F6FA', foreground='#2D2B4E',
                                font=('Segoe UI', 16, 'bold'))
            dlg_style.configure('DialogSub.TLabel', background='#F5F6FA', foreground='#8E8EA0',
                                font=('Segoe UI', 10))
            dlg_style.configure('Search.TEntry', font=('Segoe UI', 10), padding=8)
            dlg_style.configure('Status.TLabel', background='#F5F6FA', foreground='#6C63FF',
                                font=('Segoe UI', 10, 'bold'))
            dlg_style.configure('LangItem.TCheckbutton', background='white', font=('Segoe UI', 9),
                                padding=4)
            dlg_style.configure('Accent.TButton', font=('Segoe UI', 10, 'bold'), padding=10)
            dlg_style.configure('TButton', font=('Segoe UI', 9), padding=6)
            dlg_style.configure('Match.TFrame', background='#EDE7FF')
        
        main = ttk.Frame(top, style='Dialog.TFrame', padding=24)
        main.pack(fill=tk.BOTH, expand=True)
        
        title_frame = ttk.Frame(main, style='Dialog.TFrame')
        title_frame.pack(fill=tk.X, pady=(0, 20))
        ttk.Label(title_frame, text="设置默认语言", style='DialogTitle.TLabel').pack(anchor='w')
        ttk.Label(title_frame, text="选择最多 20 种常用语言，用于快速开始翻译任务", 
                  style='DialogSub.TLabel').pack(anchor='w', pady=(4, 0))
        
        search_card = ttk.Frame(main, style='DialogCard.TFrame', padding=12)
        search_card.pack(fill=tk.X, pady=(0, 12))
        
        search_row = ttk.Frame(search_card, style='DialogCard.TFrame')
        search_row.pack(fill=tk.X)
        
        ttk.Label(search_row, text="🔍", font=('Segoe UI', 14), background='white').pack(side='left')
        search_var = tk.StringVar()
        search_entry = ttk.Entry(search_row, textvariable=search_var, width=40, style='Search.TEntry')
        search_entry.pack(side='left', padx=(8, 0), fill=tk.X, expand=True)
        search_entry.insert(0, '搜索语言代码、英文或中文国家名...')
        search_entry.configure(foreground='#A5A0C0')
        
        def on_search_focus_in(e):
            if search_entry.get() == '搜索语言代码、英文或中文国家名...':
                search_entry.delete(0, tk.END)
                search_entry.configure(foreground='#2D2B4E')
        
        def on_search_focus_out(e):
            if not search_entry.get().strip():
                search_entry.insert(0, '搜索语言代码、英文或中文国家名...')
                search_entry.configure(foreground='#A5A0C0')
        
        search_entry.bind('<FocusIn>', on_search_focus_in)
        search_entry.bind('<FocusOut>', on_search_focus_out)
        
        list_card = ttk.Frame(main, style='DialogCard.TFrame', padding=16)
        list_card.pack(fill=tk.BOTH, expand=True, pady=(0, 12))
        
        canvas_frame = ttk.Frame(list_card, style='DialogCard.TFrame')
        canvas_frame.pack(fill=tk.BOTH, expand=True)
        
        canvas = tk.Canvas(canvas_frame, bg='white', highlightthickness=0)
        scrollbar = ttk.Scrollbar(canvas_frame, orient="vertical", command=canvas.yview)
        scrollable = ttk.Frame(canvas, style='DialogCard.TFrame')
        
        scrollable.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable, anchor="nw", width=600)
        canvas.configure(yscrollcommand=scrollbar.set)
        
        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        
        all_langs = [(code, info) for code, info in sorted(LANG_MAP.items())]
        lang_vars = {}
        lang_frames = {}
        
        status_bar = ttk.Frame(main, style='Dialog.TFrame')
        status_bar.pack(fill=tk.X, pady=(0, 16))
        status_label = ttk.Label(status_bar, text=f"已选择 {len(self.default_langs)}/20 种语言",
                                 style='Status.TLabel')
        status_label.pack(side='left')
        
        def update_status():
            c = sum(1 for v in lang_vars.values() if v.get())
            status_label.config(text=f"已选择 {c}/20 种语言")
        
        def on_check(code):
            v = lang_vars[code]
            if v.get() and sum(1 for x in lang_vars.values() if x.get()) > 20:
                v.set(False)
                messagebox.showwarning("提示", "最多只能选择 20 种语言", parent=top)
            update_status()
        
        for i, (code, info) in enumerate(all_langs):
            var = tk.BooleanVar(value=code in self.default_langs)
            lang_vars[code] = var
            
            item = ttk.Frame(scrollable, style='DialogCard.TFrame')
            item.pack(fill=tk.X, pady=2, padx=4)
            
            hover_bg = '#F0EEFF'
            normal_bg = 'white'
            
            def on_enter(e, frm=item):
                try: frm.configure(style='Hover.TFrame')
                except: pass
            def on_leave(e, frm=item):
                try: frm.configure(style='DialogCard.TFrame')
                except: pass
            
            item.bind('<Enter>', on_enter)
            item.bind('<Leave>', on_leave)
            
            chk = ttk.Checkbutton(item, text=f"{code}  {info['name']}", variable=var,
                                   style='LangItem.TCheckbutton',
                                   command=lambda c=code: on_check(c))
            chk.pack(anchor='w', padx=8, pady=6)
            lang_frames[code] = item
        
        def extract_cn_text(name):
            import re as _re
            cn = _re.findall(r'[\u4e00-\u9fff]+', name)
            return ' '.join(cn).lower()
        
        match_count = 0
        
        def on_search(*args):
            nonlocal match_count
            txt = search_var.get().lower().strip()
            
            if txt == '搜索语言代码、英文或中文国家名...'.lower():
                txt = ''
            
            first_match = None
            match_count = 0
            
            for code, frm in lang_frames.items():
                info = LANG_MAP[code]
                full_text = f"{code} {info['name']}".lower()
                cn_text = extract_cn_text(info['name'])
                
                if txt:
                    matched = (txt in full_text or txt in code.lower() or 
                               txt in cn_text or txt in info['name'].lower())
                    
                    if matched:
                        frm.pack(fill=tk.X, pady=2, padx=4)
                        try: frm.configure(style='Match.TFrame')
                        except: pass
                        if first_match is None:
                            first_match = frm
                        match_count += 1
                    else:
                        frm.pack_forget()
                else:
                    frm.pack(fill=tk.X, pady=2, padx=4)
                    try: frm.configure(style='DialogCard.TFrame')
                    except: pass
            
            if txt:
                status_label.config(text=f"已选择 {sum(1 for v in lang_vars.values() if v.get())}/20 种语言  |  找到 {match_count} 个匹配")
            else:
                status_label.config(text=f"已选择 {sum(1 for v in lang_vars.values() if v.get())}/20 种语言")
            
            if first_match:
                canvas.update_idletasks()
                try:
                    y = canvas.coords(canvas.find_withtag('all')[0])[1] if canvas.find_withtag('all') else 0
                    canvas.yview_moveto((frm.winfo_rooty() - scrollable.winfo_rooty() + 20) / scrollable.winfo_height())
                except Exception:
                    pass
        
        search_var.trace('w', on_search)
        
        btn_bar = ttk.Frame(main, style='Dialog.TFrame')
        btn_bar.pack(fill=tk.X)
        
        def deselect_all():
            for v in lang_vars.values(): v.set(False)
            update_status()
        
        ttk.Button(btn_bar, text="取消全选", command=deselect_all).pack(side='left')
        ttk.Separator(btn_bar, orient='vertical').pack(side='left', fill='y', padx=12, pady=4)
        
        btn_frame_right = ttk.Frame(btn_bar, style='Dialog.TFrame')
        btn_frame_right.pack(side='right')
        
        ttk.Button(btn_frame_right, text="取消", command=top.destroy).pack(side='right', padx=4)
        ok_btn = ttk.Button(btn_frame_right, text="确定保存",
                              command=lambda: self._confirm_default_langs_from_vars(top, lang_vars),
                              style='Accent.TButton')
        ok_btn.pack(side='right', padx=4)
        
        def on_mousewheel(e): canvas.yview_scroll(int(-1*(e.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", on_mousewheel)
        
        def on_close():
            canvas.unbind_all("<MouseWheel>")
            top.destroy()
        top.protocol("WM_DELETE_WINDOW", on_close)
    
    def _confirm_default_langs_from_vars(self, top, lang_check_vars):
        """
        从勾选框变量确认默认语言设置
        """
        selected_langs = [code for code, var in lang_check_vars.items() if var.get()]
        
        if selected_langs:
            self.default_langs = selected_langs[:20]
            self.save_config()
            self.show_all_langs = False
            self.lang_show_var.set(False)
            self.update_lang_display()
            messagebox.showinfo("提示", f"已设置 {len(self.default_langs)} 种默认语言")
        else:
            messagebox.showwarning("警告", "请至少选择一种语言")
        
        top.destroy()
    
    def load_translation_table(self):
        """
        加载翻译表
        从输出目录读取translation_table.json
        """
        table_path = os.path.join(self.table_dir, 'translation_table.json')
        if os.path.exists(table_path):
            try:
                with open(table_path, 'r', encoding='utf-8') as f:
                    self.translation_table = json.load(f)
                self.log(f"已加载翻译表，共 {len(self.translation_table)} 条记录")
            except Exception as e:
                self.log(f"加载翻译表失败: {e}")
    
    def save_translation_table(self):
        """
        保存翻译表
        将翻译表写入translation_table.json
        """
        table_path = os.path.join(self.table_dir, 'translation_table.json')
        try:
            with open(table_path, 'w', encoding='utf-8') as f:
                json.dump(self.translation_table, f, ensure_ascii=False, indent=2)
            self.log("翻译表已保存")
        except Exception as e:
            self.log(f"保存翻译表失败: {e}")
    
    def log(self, message):
        """
        添加日志消息
        :param message: 日志消息内容
        """
        self.log_text.insert(tk.END, f"{time.strftime('%Y-%m-%d %H:%M:%S')} - {message}\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()
    
    def extract_chinese_texts(self):
        """
        从XML或RES文件中提取中文文本
        RES文件会自动使用GZIP格式解压
        :return: 中文文本集合
        """
        all_texts = set()
        
        for source_file in self.source_files:
            try:
                # 当前source_files应为XML文本文件路径（可能为无后缀的已转换RES）
                with open(source_file, 'r', encoding='utf-8') as f:
                    content = f.read()
                
                parser = ET.XMLParser(encoding='utf-8')
                try:
                    root = ET.fromstring(content, parser=parser)
                    for elem in root.findall('.//Message'):
                        text = elem.get('Content', '')
                        if text and self.is_chinese(text):
                            all_texts.add(text)
                    self.log(f"从 {os.path.basename(source_file)} 提取到中文文本")
                except Exception as pe:
                    # 容错：XML 解析失败，尝试通过正则抽取 Message 的 Content 属性
                    self.log(f"解析 {os.path.basename(source_file)} 时 XML 解析失败: {pe}, 尝试正则容错抽取")
                    if self.logger:
                        self.logger.exception(pe)

                    try:
                        # 匹配 <Message ... Content="..." .../> 或 Content='...'
                        msgs = []
                        for m in re.finditer(r'<Message\b([^>]*)>', content, flags=re.IGNORECASE | re.DOTALL):
                            attrs = m.group(1)
                            # 尝试双引号
                            m1 = re.search(r'Content\s*=\s*"(.*?)"', attrs, flags=re.DOTALL)
                            if m1:
                                import html
                                msgs.append(html.unescape(m1.group(1)))
                                continue
                            m2 = re.search(r"Content\s*=\s*'(.*?)'", attrs, flags=re.DOTALL)
                            if m2:
                                import html
                                msgs.append(html.unescape(m2.group(1)))

                        for text in msgs:
                            if text and self.is_chinese(text):
                                all_texts.add(text)

                        self.log(f"正则方式从 {os.path.basename(source_file)} 提取到 {len(msgs)} 条 Message")
                        if len(msgs) == 0:
                            content_preview = content.strip()[:120]
                            self.log(f"警告: {os.path.basename(source_file)} 未匹配到任何Message标签，文件预览: {content_preview}")
                    except Exception as re_e:
                        self.log(f"正则抽取失败: {re_e}")
                        if self.logger:
                            self.logger.exception(re_e)
            except Exception as e:
                self.log(f"解析 {source_file} 失败: {e}")
                if self.logger:
                    self.logger.exception(f"解析失败: {source_file}")
        # 将新文本添加到翻译表
        for text in all_texts:
            if text not in self.translation_table:
                self.translation_table[text] = {lang: '' for lang in LANG_MAP.keys()}
        
        self.save_translation_table()
        return all_texts
    
    def decompress_res(self, res_file_path):
        """
        解压RES文件（GZIP格式）
        :param res_file_path: RES文件路径
        :return: 解压后的内容（字符串），失败返回None
        """
        try:
            with open(res_file_path, 'rb') as f:
                head = f.read(2)
                f.seek(0)
                if head != b'\x1f\x8b':
                    self.log(f"警告: {os.path.basename(res_file_path)} 不是有效的GZIP文件")
                    return None
                import gzip
                with gzip.GzipFile(fileobj=f) as gz:
                    data = gz.read()

            # 先尝试utf-8解码，失败则使用替代策略
            try:
                text = data.decode('utf-8')
            except UnicodeDecodeError:
                # 记录并尝试使用替代解码
                self.log(f"警告: {os.path.basename(res_file_path)} UTF-8 解码失败，使用替代解码")
                if self.logger:
                    self.logger.exception(f"UTF-8 解码失败: {res_file_path}")
                try:
                    text = data.decode('utf-8', errors='replace')
                except Exception:
                    text = data.decode('latin-1', errors='replace')

            stripped = text.strip()
            if not stripped:
                self.log(f"错误: {os.path.basename(res_file_path)} 解压后内容为空")
                return None
            if not (stripped.startswith('<?xml') or stripped.startswith('<')):
                self.log(f"错误: {os.path.basename(res_file_path)} 解压后内容不是有效XML（预览: {stripped[:80]}）")
                return None
            return text
        except Exception as e:
            self.log(f"解压错误: {e}")
            if self.logger:
                self.logger.exception(e)
            return None
    
    def is_chinese(self, text):
        """
        判断文本是否包含中文
        :param text: 待检查文本
        :return: True表示包含中文，False表示不包含
        """
        for char in text:
            if '\u4e00' <= char <= '\u9fff':
                return True
        return False
    
    def on_api_type_changed(self):
        """
        API类型切换处理
        保存用户选择的API类型
        """
        self.api_type = self.api_type_var.get()
        self.save_config()
        if self.api_type == 'baidu':
            self.log(f"当前翻译API: 百度翻译")
            if not self.baidu_app_id or not self.baidu_secret_key:
                self.log("警告: 请先配置百度翻译API密钥！")
                messagebox.showwarning("警告", "请先配置百度翻译API密钥！\n点击'API配置'按钮进行设置。")
        else:
            self.log(f"当前翻译API: Google翻译")
    
    def configure_api(self):
        """
        配置翻译API
        弹出窗口让用户输入百度翻译API的App ID和Secret Key
        """
        top = tk.Toplevel(self.root)
        top.title("API配置")
        top.geometry("500x300")
        top.transient(self.root)
        top.grab_set()
        
        main_frame = ttk.Frame(top, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 当前API类型显示
        ttk.Label(main_frame, text=f"当前使用API: {self.api_type.upper()}", font=('Arial', 11, 'bold')).pack(anchor=tk.W, pady=10)
        
        # 百度API配置
        ttk.Label(main_frame, text="百度翻译API配置:", font=('Arial', 10, 'bold')).pack(anchor=tk.W, pady=5)
        ttk.Label(main_frame, text="App ID:").pack(anchor=tk.W)
        app_id_var = tk.StringVar(value=self.baidu_app_id)
        ttk.Entry(main_frame, textvariable=app_id_var, width=50).pack(anchor=tk.W, pady=5)
        
        ttk.Label(main_frame, text="Secret Key:").pack(anchor=tk.W)
        secret_key_var = tk.StringVar(value=self.baidu_secret_key)
        ttk.Entry(main_frame, textvariable=secret_key_var, width=50, show='*').pack(anchor=tk.W, pady=5)
        
        # 说明
        ttk.Label(main_frame, text="提示: 百度翻译API需要在百度翻译开放平台申请", 
                  font=('Arial', 9), foreground='gray').pack(anchor=tk.W, pady=10)
        
        # 按钮
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(anchor=tk.E, pady=20)
        
        def save_and_close():
            self.baidu_app_id = app_id_var.get().strip()
            self.baidu_secret_key = secret_key_var.get().strip()
            self.save_config()
            messagebox.showinfo("提示", "API配置已保存！")
            top.destroy()
        
        ttk.Button(btn_frame, text="保存", command=save_and_close).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="取消", command=top.destroy).pack(side=tk.LEFT, padx=5)
    
    def baidu_translate(self, text, target_lang):
        """
        使用百度翻译API翻译文本
        :param text: 待翻译文本
        :param target_lang: 目标语言代码
        :return: 翻译后的文本
        """
        if not self.baidu_app_id or not self.baidu_secret_key:
            self.log("错误: 百度API密钥未配置！")
            return text
        
        # 百度翻译API支持的语言列表
        supported_langs = {
            'auto', 'zh', 'en', 'yue', 'wyw', 'ja', 'ko', 'fr', 'de', 'es', 'pt', 'pt-PT',
            'vi', 'id', 'th', 'ru', 'ar', 'hi', 'bn', 'pa', 'gu', 'or', 'ta', 'te', 'kn',
            'ml', 'mr', 'sa', 'yi', 'hi-IN', 'ne', 'my', 'si', 'km', 'lo', 'bo', 'ug',
            'ca', 'eu', 'gl', 'it', 'el', 'hu', 'cs', 'sk', 'sl', 'pl', 'ro', 'bg',
            'sr', 'hr', 'lv', 'lt', 'et', 'fi', 'sv', 'da', 'no', 'is', 'mt', 'tr',
            'hi-Latn', 'ur', 'fa', 'sq', 'bs', 'ms', 'ms-Arab', 'tl', 'ht', 'sw',
            'xh', 'zu', 'af', 'zu', 'xh', 'st', 'tn', 'ts', 'ss', 'nr', 've', 'wo',
            'rm', 'ga', 'gd', 'cy', 'kw', 'br', 'gv', 'gd', 'ga', 'eu', 'gl', 'ast',
            'ca', 'oc', 'fr-CA', 'it-CH', 'de-CH', 'gsw', 'als', 'lb', 'wa', 'nl',
            'fy', 'pap', 'sv-SE', 'sv-FI', 'nb', 'nn', 'da', 'is', 'fo', 'et', 'lv',
            'lt', 'pl', 'cs', 'sk', 'sl', 'hr', 'sr', 'bg', 'mk', 'sq', 'mt', 'tr',
            'az', 'uz', 'kk', 'ky', 'tg', 'tt', 'ba', 'cv', 'ce', 'yi', 'he', 'ar',
            'fa', 'ur', 'ps', 'sd', 'ku', 'ckb', 'syc', 'dv', 'hi', 'bn', 'as',
            'gu', 'or', 'ta', 'te', 'kn', 'ml', 'pa', 'mr', 'sa', 'kok', 'mai',
            'ne', 'si', 'my', 'km', 'lo', 'th', 'ja', 'ko', 'zh', 'zh-TW', 'zh-HK',
            'zh-CN', 'yue', 'wyw', 'vi', 'id', 'tl', 'ms', 'en', 'de', 'fr', 'es',
            'pt', 'ru', 'it', 'pl', 'ro', 'bg', 'cs', 'da', 'nl', 'fi', 'el', 'hu',
            'no', 'pt-PT', 'sv', 'tr', 'ar', 'hi', 'ja', 'ko', 'th', 'vi', 'zh',
            'zh-TW', 'en', 'de', 'fr', 'es', 'pt', 'ru', 'it', 'pl', 'ro', 'bg',
            'cs', 'da', 'nl', 'fi', 'el', 'hu', 'no', 'pt-PT', 'sv', 'tr', 'ar',
            'hi', 'ja', 'ko', 'th', 'vi', 'zh', 'zh-TW', 'en', 'de', 'fr', 'es',
            'pt', 'ru', 'it', 'pl', 'ro', 'bg', 'cs', 'da', 'nl', 'fi', 'el', 'hu',
            'no', 'pt-PT', 'sv', 'tr'
        }
        
        # 调试日志
        self.log(f"百度翻译: target_lang={target_lang}")
        
        # 百度翻译API语言代码映射表
        baidu_lang_map = {
            # 中文变体
            'zh-CN': 'zh', 'zh-TW': 'cht', 'zh-HK': 'zh', 'zh-SG': 'zh', 'zh-MO': 'zh',
            # 阿拉伯语变体
            'ar-SA': 'ar', 'ar-IQ': 'ar', 'ar-EG': 'ar', 'ar-LY': 'ar', 'ar-DZ': 'ar',
            'ar-MA': 'ar', 'ar-TN': 'ar', 'ar-OM': 'ar', 'ar-YE': 'ar', 'ar-SY': 'ar',
            'ar-JO': 'ar', 'ar-LB': 'ar', 'ar-KW': 'ar', 'ar-AE': 'ar', 'ar-BH': 'ar',
            'ar-QA': 'ar',
            # 德语变体
            'de-DE': 'de', 'de-CH': 'de', 'de-AT': 'de', 'de-LU': 'de', 'de-LI': 'de',
            # 英语变体
            'en-US': 'en', 'en-GB': 'en', 'en-AU': 'en', 'en-CA': 'en', 'en-NZ': 'en',
            'en-IE': 'en', 'en-ZA': 'en',
            # 西班牙语变体
            'es-ES': 'es', 'es-MX': 'es', 'es-GT': 'es', 'es-CR': 'es', 'es-PA': 'es',
            'es-DO': 'es', 'es-VE': 'es', 'es-CO': 'es', 'es-PE': 'es', 'es-AR': 'es',
            'es-EC': 'es', 'es-CL': 'es', 'es-UY': 'es', 'es-PY': 'es', 'es-BO': 'es',
            'es-SV': 'es', 'es-HN': 'es', 'es-NI': 'es', 'es-PR': 'es',
            # 法语变体
            'fr-FR': 'fr', 'fr-BE': 'fr', 'fr-CA': 'fr', 'fr-CH': 'fr', 'fr-LU': 'fr',
            'fr-MC': 'fr',
            # 意大利语变体
            'it-IT': 'it', 'it-CH': 'it',
            # 荷兰语变体
            'nl-NL': 'nl', 'nl-BE': 'nl',
            # 葡萄牙语变体
            'pt-PT': 'pt', 'pt-BR': 'pt',
            # 塞尔维亚语变体
            'sr-Latn': 'sr', 'sr-Cyrl': 'sr',
            # 瑞典语变体
            'sv-SE': 'sv', 'sv-FI': 'sv',
            # 乌兹别克语变体
            'uz-Latn': 'uz',
            # 阿塞拜疆语变体
            'az-Latn': 'az',
            # 其他语言
            'ca': 'ca', 'bg': 'bg', 'cs': 'cs', 'da': 'da', 'el': 'el', 'eo': 'eo',
            'et': 'et', 'fi': 'fi', 'he': 'he', 'hu': 'hu', 'is': 'is', 'ja': 'ja',
            'ko': 'ko', 'lv': 'lv', 'lt': 'lt', 'mk': 'mk', 'no': 'no', 'nn': 'no',
            'nb': 'no', 'pl': 'pl', 'ro': 'ro', 'ru': 'ru', 'hr': 'hr', 'sk': 'sk',
            'sl': 'sl', 'sq': 'sq', 'th': 'th', 'tr': 'tr', 'ur-PK': 'ur', 'id': 'id',
            'uk': 'uk', 'be': 'be', 'fa': 'fa', 'vi': 'vi', 'hy': 'hy', 'eu': 'eu',
            'af': 'af', 'ka': 'ka', 'fo': 'fo', 'hi': 'hi', 'ms-MY': 'ms', 'ms-BN': 'ms',
            'kk': 'kk', 'sw': 'sw', 'tt': 'tt', 'bn': 'bn', 'pa': 'pa', 'gu': 'gu',
            'or': 'or', 'ta': 'ta', 'te': 'te', 'kn': 'kn', 'ml': 'ml', 'as': 'as',
            'mr': 'mr', 'sa': 'sa', 'kok': 'kok'
        }
        
        # 获取百度语言代码
        if target_lang in baidu_lang_map:
            lang_code = baidu_lang_map[target_lang]
            self.log(f"百度翻译: 映射成功 {target_lang} -> {lang_code}")
        else:
            # 尝试仅使用语言代码部分
            base_lang = target_lang.split('-')[0] if '-' in target_lang else target_lang
            if base_lang in baidu_lang_map:
                lang_code = baidu_lang_map[base_lang]
                self.log(f"百度翻译: 使用基础语言代码 {target_lang} -> {base_lang} -> {lang_code}")
            else:
                self.log(f"警告: 百度翻译不支持语言 '{target_lang}'，使用英语")
                lang_code = 'en'
        
        max_retries = 3
        for attempt in range(max_retries):
            try:
                salt = str(uuid.uuid4())[:10]
                sign = self.baidu_app_id + text + salt + self.baidu_secret_key
                sign = hashlib.md5(sign.encode('utf-8')).hexdigest()
                
                url = 'https://fanyi-api.baidu.com/api/trans/vip/translate'
                data = {
                    'q': text[:50] + '...' if len(text) > 50 else text,
                    'from': 'auto',
                    'to': lang_code,
                    'appid': self.baidu_app_id,
                    'salt': salt,
                    'sign': sign
                }
                
                self.log(f"百度翻译请求: from=auto, to={lang_code}, text_len={len(text)}")
                
                response = requests.post(url, data=data, timeout=15)
                result = response.json()
                
                self.log(f"百度翻译响应: {result}")
                
                if 'trans_result' in result and result['trans_result']:
                    return result['trans_result'][0]['dst']
                else:
                    if 'error_code' in result:
                        error_code = result['error_code']
                        error_msg = result.get('error_msg', '未知错误')
                        self.log(f"百度翻译API错误[{error_code}]: {error_msg}")
                        # 如果是语言方向不支持，提示用户检查账户权限
                        if error_code == '58001':
                            self.log(f"提示: 请检查百度翻译API账户是否开通了 {lang_code} 语言的翻译服务")
                    return text
            except Exception as e:
                self.log(f"百度翻译异常: {str(e)}")
                if attempt < max_retries - 1:
                    time.sleep(2)
                    continue
                self.log(f"百度翻译失败: {text} -> {target_lang}")
                return text
        
        return text
    
    def translate_text(self, text, target_lang):
        """
        使用选择的翻译API翻译文本
        :param text: 待翻译文本
        :param target_lang: 目标语言代码
        :return: 翻译后的文本
        """
        text_strip = text.strip()
        if not text_strip:
            return text
        if text_strip in BRAND_NAMES:
            return text
        
        if self.api_type == 'baidu':
            return self.baidu_translate(text, target_lang)
        
        max_retries = 3
        for attempt in range(max_retries):
            try:
                url = f'https://translate.googleapis.com/translate_a/single?client=gtx&sl=zh-CN&tl={target_lang}&dt=t&q={quote(text)}'
                headers = {'User-Agent': 'Mozilla/5.0'}
                response = requests.get(url, headers=headers, timeout=15)
                response.raise_for_status()
                result = response.json()
                translated = ''.join([item[0] for item in result[0] if item[0]])
                return translated
            except Exception as e:
                if attempt < max_retries - 1:
                    time.sleep(3)
                    continue
                self.log(f"翻译失败: {text} -> {target_lang}")
                return text
    
    def start_translation(self):
        """
        开始翻译流程
        1. 检查输入
        2. 提取中文文本
        3. 执行翻译
        4. 更新进度
        """
        # 检查输入
        if not self.source_files:
            msg = "请先添加文件或文件夹，或点击DiskC工作流"
            if self.headless:
                self.log(f"错误: {msg}")
                if self.logger:
                    self.logger.error(msg)
                return
            else:
                messagebox.showwarning("警告", msg)
                return

        # 获取选中的语言（GUI模式）或保持已有设置（headless可在外部设置）
        if not self.headless:
            self.selected_langs = [lang for lang, var in self.lang_vars.items() if var.get()]

        if not self.selected_langs:
            msg = "请至少选择一种目标语言"
            if self.headless:
                self.log(f"错误: {msg}")
                if self.logger:
                    self.logger.error(msg)
                return
            else:
                messagebox.showwarning("警告", msg)
                return
        
        # 初始化状态
        self.start_btn.config(state='disabled')
        self.confirm_btn.config(state='disabled')
        self.translation_complete = False
        self.progress_var.set(0)
        self.progress_label.config(text="0%")
        self.log("开始提取中文文本...")
        
        # 提取中文文本
        texts = self.extract_chinese_texts()
        self.log(f"共需处理 {len(texts)} 条中文文本")
        
        # 计算需要翻译的数量
        total_needed = 0
        for text, langs in self.translation_table.items():
            for lang in self.selected_langs:
                if lang not in langs:
                    langs[lang] = ''
                if not langs[lang] or langs[lang] == text:
                    total_needed += 1
        
        # 如果没有需要翻译的内容
        if total_needed == 0:
            self.log("所有翻译已完成")
            self.translation_complete = True
            self.confirm_btn.config(state='normal')
            self.start_btn.config(state='normal')
            return
        
        self.log(f"需要翻译: {total_needed} 条")
        
        # 创建翻译线程
        def translate_thread():
            from concurrent.futures import ThreadPoolExecutor, as_completed

            texts_to_translate = []
            for text in self.translation_table.keys():
                for lang in self.selected_langs:
                    if lang not in self.translation_table[text]:
                        self.translation_table[text][lang] = ''
                    if not self.translation_table[text][lang] or self.translation_table[text][lang] == text:
                        texts_to_translate.append((text, lang))

            completed = 0
            lock = threading.Lock()

            def translate_one(item):
                text, lang = item
                result = self.translate_text(text, LANG_MAP[lang]['code'])
                return text, lang, result

            with ThreadPoolExecutor(max_workers=5) as executor:
                futures = {executor.submit(translate_one, item): item for item in texts_to_translate}
                for future in as_completed(futures):
                    text, lang, result = future.result()
                    current_completed = 0
                    current_progress = 0.0
                    should_save = False
                    with lock:
                        self.translation_table[text][lang] = result
                        completed += 1
                        current_completed = completed
                        current_progress = round(completed / total_needed * 100, 1)
                        if completed % 50 == 0:
                            should_save = True

                    if should_save:
                        self.save_translation_table()

                    self.root.after(0, lambda p=current_progress, c=current_completed: self._update_progress(p, c))
                    self.log(f"进度: {current_progress}% ({current_completed}/{total_needed})")

            self.log("执行智能缩写...")
            self.abbreviate_translations()
            self.save_translation_table()
            self.root.after(0, lambda: self._update_progress(100, total_needed))
            self.log("翻译完成！")
            self.translation_complete = True
            self.root.after(0, lambda: self.confirm_btn.config(state='normal'))
            self.root.after(0, lambda: self.start_btn.config(state='normal'))
        
        # 启动翻译线程
        threading.Thread(target=translate_thread, daemon=True).start()
    
    def abbreviate_translations(self):
        import re

        max_lengths = {
            'default': 40,
            'JPN': 22, 'KOR': 22,
            'GER': 35, 'DES': 35, 'DEA': 35, 'DEL': 35, 'DEC': 35,
            'FRA': 32, 'FRB': 32, 'FRC': 32, 'FRS': 32, 'FRL': 32, 'FRM': 32,
            'ITA': 30, 'ITS': 30,
            'ESP': 28, 'ESM': 28, 'ESG': 28, 'ESC': 28, 'ESA': 28, 'ESD': 28,
            'ESV': 28, 'ESO': 28, 'ESR': 28, 'ESS': 28, 'ESF': 28, 'ESL': 28,
            'ESY': 28, 'ESB': 28, 'ESE': 28, 'ESH': 28, 'ESN': 28, 'ESU': 28, 'ESP2': 28,
            'PTG': 30, 'PTB': 30,
            'RUS': 25,
            'ARA': 40, 'ARL': 40, 'ARG': 40, 'ARM': 40, 'ART': 40,
            'ARO': 40, 'ARY': 40, 'ARS': 40, 'ARJ': 40, 'ARB': 40,
            'ARK': 40, 'ARU': 40, 'ARH': 40, 'ARQ': 40,
        }

        CN_FUNCTION_WORDS = set(
            '的 是 在 和 与 及 了 有 我 你 他 她 它 能 可 将 已 被 从 到 对 为 于 以 此 其 该 这 那 些 等 或 但 如 若 则 因 所 并 且 而 又 也 更 最 很 太 较 稍 略 仅 只 均 皆 各 每 任 某 另 再 还 仍 尚 须 应 需 要 会 得 让 把 比 按 照 根 据 由 向 往 朝 沿 顺 随 跟 同 非 无 不 没 未 否 勿 别 莫 休 罢 着 过 起 来 去 出 入 上 下 左 右 前 后 内 外 中 间 里 边 旁 侧 底 顶 头 尾 首 末 始 终 全 整 部 分 段 节 项 条 类 种 型 式 样 法 方 式 途 径 路 线 面 体 点 位 置 处 所 场 地 域 区 范 围 界 限 度 量 数 值 参 据 信 息 号 码 标 志 记 符 名 称 编 序 代 码 键 值 组 列 表 格 框 窗 页 屏 幕 板 卡 盘 钮 键 开 关 启 停 复 重 设 置 配 选 确 取 消 删 除 增 加 修 改 更 新 换 替 转 变 化 显 示 隐 藏 查 找 搜 索 读 写 存 取 载 卸 装 连 接 断 开 通 发 送 收 受 传 输 播 放 录 记 打 印 预 览 扫 描 检 测 试 验 证 核 校 正 调 整 优 化 改 善 提 升 降 低 减 多 扩 大 缩 拉 伸 压 移 动 拖 拽 滚 翻 页 跳 转 返 回 退 进'
        )

        EN_STOP_WORDS = set(
            'a an the and or but if is are was were be been being have has had do does did will would shall should can could may might must need want like love hate think know believe see hear feel make get give take come go use find tell ask say speak talk walk run work play start stop begin end open close save load create delete add remove set get put show hide enable disable select choose click press enter exit cancel ok yes no true false on off in at to from by for of with about above after before between under over through into onto upon within without during since until while as than up down left right back front near far all any each every some many few more most less much such other another same different new old first last next previous current default custom user system file folder directory window page menu button label text input output setting option mode state status type kind form format size color style value name number id index key code data info message error warning success fail pass result total count sum average min max low high auto manual public private local global remote main sub help tip note log view edit copy cut paste paste undo redo clear reset refresh reload search sort filter group merge split join connect disconnect attach detach lock unlock freeze thaw zoom in out expand collapse'
        )

        JP_PARTICLES = set('の は が を に へ と で から まで も など しか も か なら けれど けど し て に な だ です ます た ない ぬ ね よ わ ぁ い う え お'.split())

        KR_PARTICLES = set('의 에서 은 는 를 을 가 과 와 도 부터 까지 조차 만큼 처럼'.split())

        def abbreviate_cn(text, limit):
            chars = list(text)
            kept = [c for c in chars if c not in CN_FUNCTION_WORDS]
            result = ''.join(kept)
            if len(result) <= limit:
                return result
            head = result[:limit//2]
            tail = result[-(limit//2-2):] if len(result) > limit else ''
            return f"{head}…{tail}" if tail else f"{result[:limit-1]}…"

        def abbreviate_en(text, limit):
            words = text.split()
            if len(words) <= 2:
                return text[:limit] + ('...' if len(text) > limit else '')
            kept = []
            for i, w in enumerate(words):
                w_lower = w.lower().rstrip('.,;:!?')
                if i == 0:
                    kept.append(w)
                elif w_lower in EN_STOP_WORDS or len(w_lower) <= 1:
                    continue
                elif len(w) > 7:
                    kept.append(w[:5] + '.')
                else:
                    kept.append(w)
                if len(' '.join(kept)) >= limit - 3:
                    break
            result = ' '.join(kept)
            if len(result) <= limit:
                return result
            return f"{result[:limit//2]}…{result[-(limit//2-2):]}" if len(result) > limit else f"{result[:limit-1]}…"

        def abbreviate_jp(text, limit):
            chars = list(text)
            kept = [c for c in chars if c not in JP_PARTICLES]
            result = ''.join(kept)
            if len(result) <= limit:
                return result
            return f"{result[:limit-1]}…"

        def abbreviate_kr(text, limit):
            syllables = re.findall(r'[\uAC00-\uD7AF\u1100-\u11FF\u3130-\u318F]|[^\s]', text)
            result = ''.join(syllables)
            if len(result) <= limit:
                return result
            return f"{result[:limit-1]}…"

        def abbreviate_default(text, limit):
            words = text.split()
            if len(words) <= 2:
                return text[:limit] + ('...' if len(text) > limit else '')
            kept = [words[0]]
            for w in words[1:]:
                w_clean = w.lower().rstrip('.,;:!?')
                if w_clean in EN_STOP_WORDS or len(w_clean) <= 1:
                    continue
                elif len(w) > 6:
                    kept.append(w[:4] + '.')
                else:
                    kept.append(w)
                if len(' '.join(kept)) >= limit - 3:
                    break
            result = ' '.join(kept)
            return result if len(result) <= limit else f"{result[:limit-1]}…"

        abbreviated_count = 0
        total_to_check = sum(
            1 for text in self.translation_table.keys()
            for lang in self.selected_langs
            if lang in self.translation_table[text] and self.translation_table[text][lang]
            and len(self.translation_table[text][lang]) > max_lengths.get(lang, max_lengths['default'])
            and lang not in ('CHT', 'CHS', 'ZHH', 'ZHI', 'ZHM')
        )
        checked = 0
        for text in self.translation_table.keys():
            for lang in self.selected_langs:
                if lang not in self.translation_table[text] or not self.translation_table[text][lang]:
                    continue
                original = self.translation_table[text][lang]
                max_len = max_lengths.get(lang, max_lengths['default'])
                if len(original) <= max_len:
                    continue
                if lang in ('CHT', 'CHS', 'ZHH', 'ZHI', 'ZHM'):
                    continue
                elif lang == 'JPN':
                    new_val = abbreviate_jp(original, max_len)
                elif lang == 'KOR':
                    new_val = abbreviate_kr(original, max_len)
                elif lang in ('ENG', 'USA', 'ENA', 'ENC', 'ENZ', 'ENI', 'ENS'):
                    new_val = abbreviate_en(original, max_len)
                elif lang in ('GER', 'DES', 'DEA', 'DEL', 'DEC'):
                    new_val = abbreviate_en(original, max_len)
                elif lang in ('FRA', 'FRB', 'FRC', 'FRS', 'FRL', 'FRM'):
                    new_val = abbreviate_en(original, max_len)
                elif lang in ('ITA', 'ITS'):
                    new_val = abbreviate_en(original, max_len)
                elif lang in ('ESP', 'ESM', 'ESG', 'ESC', 'ESA', 'ESD',
                              'ESV', 'ESO', 'ESR', 'ESS', 'ESF', 'ESL',
                              'ESY', 'ESB', 'ESE', 'ESH', 'ESN', 'ESU', 'ESP2'):
                    new_val = abbreviate_en(original, max_len)
                elif lang in ('PTG', 'PTB'):
                    new_val = abbreviate_en(original, max_len)
                elif lang == 'RUS':
                    new_val = abbreviate_default(original, max_len)
                elif lang in ('ARA', 'ARL', 'ARG', 'ARM', 'ART', 'ARO',
                              'ARY', 'ARS', 'ARJ', 'ARB', 'ARK', 'ARU', 'ARH', 'ARQ'):
                    new_val = abbreviate_default(original, max_len)
                else:
                    new_val = abbreviate_default(original, max_len)
                    new_val = abbreviate_default(original, max_len)
                self.translation_table[text][lang] = new_val
                abbreviated_count += 1
                checked += 1
                if checked % 200 == 0:
                    pct = round(checked / max(total_to_check, 1) * 100, 1)
                    self.log(f"智能缩写: {pct}% ({checked}/{total_to_check})")

        self.log(f"智能缩写完成，共缩写 {abbreviated_count} 条翻译")
    
    def _update_progress(self, progress, count):
        self.progress_var.set(progress)
        self.progress_label.config(text=f"{progress}%")
        if hasattr(self, 'progress_label') and progress > 0:
            master = self.progress_label.master
            for child in master.winfo_children():
                if isinstance(child, ttk.Label) and child.cget('text').startswith(' '):
                    child.config(text="  处理中..." if progress < 100 else "  完成!")
                    break
    
    def confirm_and_generate(self):
        """
        确认翻译并生成XML文件
        需要用户确认后才执行生成操作
        """
        if not self.translation_complete:
            messagebox.showwarning("警告", "请先完成翻译")
            return
        
        result = messagebox.askyesno("确认生成", "翻译已完成，是否确认生成XML文件？")
        if result:
            self.generate_xml_files()
    
    def generate_xml_files(self):
        """
        生成各语言版本的XML文件
        将翻译结果写入对应语言文件夹
        """
        if not self.source_files:
            msg = "请先选择XML源文件"
            if self.headless:
                self.log(f"错误: {msg}")
                if self.logger:
                    self.logger.error(msg)
                return
            else:
                messagebox.showwarning("警告", msg)
                return

        if not self.selected_langs:
            msg = "请至少选择一种目标语言"
            if self.headless:
                self.log(f"错误: {msg}")
                if self.logger:
                    self.logger.error(msg)
                return
            else:
                messagebox.showwarning("警告", msg)
                return
        
        self.log("开始生成XML文件...")
        
        for xml_file in self.source_files:
            try:
                with open(xml_file, 'r', encoding='utf-8') as f:
                    content = f.read()

                messages = []
                # 统一使用正则提取，保留原始Message元素的完整文本（含所有属性如id/ID/Content等）
                try:
                    for m in re.finditer(r'<Message\b[^>]*/>', content, flags=re.IGNORECASE | re.DOTALL):
                        raw_elem = m.group(0)
                        id_m = (re.search(r'\bID\s*=\s*"(.*?)"', raw_elem) or
                                re.search(r"\bID\s*=\s*'(.*?)'", raw_elem) or
                                re.search(r'\bid\s*=\s*"(.*?)"', raw_elem) or
                                re.search(r"\bid\s*=\s*'(.*?)'", raw_elem))
                        msg_id = id_m.group(1) if id_m else ''
                        cont_m = (re.search(r'Content\s*=\s*"(.*?)"', raw_elem) or
                                  re.search(r"Content\s*=\s*'(.*?)'", raw_elem))
                        original_content = cont_m.group(1) if cont_m else ''
                        import html
                        original_content = html.unescape(original_content)
                        messages.append((msg_id, original_content, raw_elem))
                    self.log(f"从 {os.path.basename(xml_file)} 提取 {len(messages)} 条 Message")
                    if len(messages) == 0:
                        content_preview = content.strip()[:120]
                        self.log(f"警告: {os.path.basename(xml_file)} 未匹配到任何Message标签，文件预览: {content_preview}")
                except Exception as ex_e:
                    self.log(f"Message提取失败: {ex_e}")
                    if self.logger:
                        self.logger.exception(ex_e)

                self.log(f"从 {os.path.basename(xml_file)} 提取 {len(messages)} 条Message（含非中文），准备生成翻译文件")

                file_name = os.path.basename(xml_file)
                if file_name.endswith('.xml'):
                    base_name = file_name[:-4]
                else:
                    base_name = file_name

                is_diskc_res = self.diskc_mode and xml_file == self.diskc_res_source
                is_diskc_xml = self.diskc_mode and xml_file in self.diskc_xml_map

                # 为每种选中的语言生成XML文件
                for lang in self.selected_langs:
                    if is_diskc_res:
                        language_dir = os.path.join(self.diskc_root, 'OpenCNC', 'Bin', 'Language')
                        os.makedirs(language_dir, exist_ok=True)
                        xml_output_path = os.path.join(language_dir, lang)
                        pack_after_write = True
                    elif is_diskc_xml:
                        rel_path = self.diskc_xml_map[xml_file]
                        string_dir = os.path.join(self.diskc_root, 'OpenCnc Shared', 'OCRes', lang, 'String')
                        os.makedirs(os.path.join(string_dir, os.path.dirname(rel_path)), exist_ok=True)
                        xml_output_path = os.path.join(string_dir, rel_path)
                        pack_after_write = False
                    else:
                        xml_output_path = os.path.join(self.output_dir, lang, 'String', base_name + '.xml')
                        os.makedirs(os.path.dirname(xml_output_path), exist_ok=True)
                        pack_after_write = self.pack_var.get()

                    lines = []
                    lines.append('<?xml version="1.0" encoding="utf-8"?>')
                    lines.append('<ResMap>')

                    if pack_after_write:
                        lang_identity = LANG_NATIVE_NAME.get(lang, lang)
                        lang_identity = lang_identity.replace('\n', '&#xA;').replace('"', '&quot;')
                        lines.append(f'  <Message ID="{lang}" Content="{lang_identity}" />')

                    for msg_id, original_content, raw_elem in messages:
                        if original_content in self.translation_table and lang in self.translation_table[original_content]:
                            translated = self.translation_table[original_content][lang]
                            content_val = translated if translated and translated != original_content else original_content
                        else:
                            content_val = original_content

                        content_val = content_val.replace('\n', '&#xA;')
                        content_val = content_val.replace('"', '&quot;')
                        # 在原始元素文本中仅替换Content值，保留id/ID等其他所有属性
                        new_elem = re.sub(
                            r'Content\s*=\s*"[^"]*"',
                            f'Content="{content_val}"',
                            raw_elem
                        )
                        lines.append(f'  {new_elem.strip()}')

                    lines.append('</ResMap>')

                    try:
                        with open(xml_output_path, 'w', encoding='utf-8') as f_out:
                            f_out.write('\n'.join(lines))
                        self.log(f"生成 {lang}: {os.path.basename(xml_output_path)} ({len(messages)} 条Message)")
                    except Exception as write_e:
                        self.log(f"写入文件失败: {write_e}")
                        if self.logger:
                            self.logger.exception(write_e)

                    if pack_after_write:
                        self.pack_to_res(xml_output_path)
            except Exception as e:
                self.log(f"生成XML失败: {e}")
                if self.logger:
                    self.logger.exception(e)
        
        self.log("所有XML文件已生成完成！")
        
        self._cleanup_temp_files()
        
        if self.diskc_mode:
            messagebox.showinfo("完成", f"DiskC工作流处理完成！\n"
                                f"RES文件已打包至: OpenCNC/Bin/Language/\n"
                                f"XML文件已输出至: OpenCnc Shared/OCRes/\n"
                                f"临时文件已清理")
        elif self.pack_var.get():
            messagebox.showinfo("完成", "XML文件已成功生成并打包为.res文件！\n临时文件已清理")
        else:
            messagebox.showinfo("完成", "XML文件已成功生成！\n临时文件已清理")
    
    def _cleanup_temp_files(self):
        import shutil
        
        self.log("开始清理临时文件...")
        
        conv_dir = os.path.join(self.output_dir, '_res_converted')
        if os.path.exists(conv_dir):
            try:
                shutil.rmtree(conv_dir)
                self.log(f"已删除: {conv_dir}")
            except Exception as e:
                self.log(f"删除失败: {conv_dir} -> {e}")
        
        if self.diskc_mode and self.diskc_root:
            lang_dir = os.path.join(self.diskc_root, 'OpenCNC', 'Bin', 'Language')
            if os.path.isdir(lang_dir):
                deleted_count = 0
                for f in os.listdir(lang_dir):
                    fp = os.path.join(lang_dir, f)
                    if os.path.isfile(fp) and not os.path.splitext(fp)[1]:
                        try:
                            os.remove(fp)
                            deleted_count += 1
                        except Exception as e:
                            self.log(f"删除失败: {fp} -> {e}")
                if deleted_count > 0:
                    self.log(f"已删除 Language 目录下 {deleted_count} 个无后缀文件")
        
        self.log("临时文件清理完成")
    
    def pack_to_res(self, xml_file_path):
        """
        将XML文件打包为GZIP格式的.res文件
        """
        try:
            with open(xml_file_path, 'rb') as f_in:
                content = f_in.read()
            
            import zlib
            crc = zlib.crc32(content) & 0xffffffff
            compressed = zlib.compress(content, 9)
            
            # 构造GZIP文件头（确保FLG=0x00）
            gzip_header = b'\x1f\x8b\x08\x00'  # ID1, ID2, CM, FLG
            gzip_header += b'\x00\x00\x00\x00'  # MTIME (0)
            gzip_header += b'\x00'              # XFL (0)
            gzip_header += b'\xff'              # OS (255 = unknown)
            
            # 生成.res文件路径
            if xml_file_path.endswith('.xml'):
                res_path = xml_file_path[:-4] + '.res'
            else:
                res_path = xml_file_path + '.res'
            
            # 写入文件
            with open(res_path, 'wb') as f_out:
                f_out.write(gzip_header)
                f_out.write(compressed[2:-4])  # 移除zlib头和adler32
                # 添加GZIP尾部（CRC32和原始大小）
                f_out.write(crc.to_bytes(4, 'little'))
                f_out.write(len(content).to_bytes(4, 'little'))
            
            self.log(f"打包: {res_path}")
        except Exception as e:
            self.log(f"打包失败: {e}")
    
    def view_translation_table(self):
        """
        打开翻译表文件
        """
        table_path = os.path.join(self.table_dir, 'translation_table.json')
        if os.path.exists(table_path):
            os.startfile(table_path)
        else:
            messagebox.showinfo("提示", "翻译表文件不存在")
    
    def export_log(self):
        """
        导出日志文件
        """
        log_content = self.log_text.get('1.0', tk.END)
        log_path = os.path.join(self.output_dir, 'translation_log.txt')
        with open(log_path, 'w', encoding='utf-8') as f:
            f.write(log_content)
        messagebox.showinfo("提示", f"日志已导出到: {log_path}")
    
    def export_to_excel(self):
        """
        导出翻译表到Excel文件
        方便人工翻译和编辑
        """
        if not self.translation_table:
            messagebox.showwarning("警告", "翻译表为空，请先进行翻译")
            return
        
        # 动态获取当前勾选的语言
        current_selected_langs = [lang for lang, var in self.lang_vars.items() if var.get()]
        if not current_selected_langs:
            messagebox.showwarning("警告", "请先选择目标语言")
            return
        
        file_path = filedialog.asksaveasfilename(
            title="导出Excel文件",
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx")]
        )
        
        if not file_path:
            return
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "翻译表"
            
            # 表头：原文 + 各语言（显示语言缩写和名称）
            headers = ["原文"]
            for lang in current_selected_langs:
                lang_name = LANG_MAP.get(lang, {}).get('name', lang)
                headers.append(f"{lang} - {lang_name}")
            ws.append(headers)
            
            # 填充数据
            for original_text, translations in self.translation_table.items():
                row = [original_text]
                for lang in current_selected_langs:
                    row.append(translations.get(lang, ''))
                ws.append(row)
            
            # 自动调整列宽
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column].width = adjusted_width
            
            wb.save(file_path)
            self.log(f"翻译表已导出到Excel: {file_path}")
            messagebox.showinfo("完成", f"翻译表已成功导出到:\n{file_path}")
        except Exception as e:
            messagebox.showerror("错误", f"导出Excel失败: {str(e)}")
    
    def import_from_excel(self):
        """
        从Excel文件导入翻译结果
        支持人工编辑后重新导入
        """
        file_path = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[("Excel文件", "*.xlsx")]
        )
        
        if not file_path:
            return
        
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            
            # 获取表头
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)
            
            if headers[0] != "原文":
                messagebox.showwarning("警告", "Excel文件格式不正确，第一列必须是'原文'")
                return
            
            # 获取语言列表（解析列标题，提取语言缩写）
            import_langs = []
            for header in headers[1:]:
                if header:
                    # 解析格式: "语言缩写 - 语言名称"
                    parts = str(header).split(' - ', 1)
                    lang_code = parts[0].strip() if parts else str(header).strip()
                    import_langs.append(lang_code)
            
            # 导入数据
            updated_count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                
                original_text = row[0]
                
                # 创建或更新翻译表条目
                if original_text not in self.translation_table:
                    self.translation_table[original_text] = {}
                
                # 更新各语言翻译
                for i, lang in enumerate(import_langs):
                    if i + 1 < len(row) and row[i + 1]:
                        old_value = self.translation_table[original_text].get(lang, '')
                        if old_value != row[i + 1]:
                            self.translation_table[original_text][lang] = row[i + 1]
                            updated_count += 1
            
            # 保存翻译表
            self.save_translation_table()
            
            self.log(f"从Excel导入完成，共更新 {updated_count} 条翻译")
            messagebox.showinfo("完成", f"成功导入 {updated_count} 条翻译")
        except Exception as e:
            messagebox.showerror("错误", f"导入Excel失败: {str(e)}")


if __name__ == '__main__':
    """
    程序入口
    创建主窗口并启动应用
    """
    parser = argparse.ArgumentParser(description='自动化翻译处理工具（GUI/命令行）')
    parser.add_argument('--input', '-i', help='输入文件或文件夹路径（支持XML或RES）')
    parser.add_argument('--mode', choices=['batch', 'single'], default='batch', help='处理模式: batch=文件夹, single=单文件')
    parser.add_argument('--output', '-o', help='输出目录（默认为脚本目录）')
    parser.add_argument('--langs', help='逗号分隔的目标语言缩写（例如: CHS,ENG）')
    parser.add_argument('--pack', action='store_true', help='处理后是否打包为.res')
    parser.add_argument('--diskc', help='DiskC工作流模式: 指定DiskC根目录路径')
    parser.add_argument('--log', help='将详细日志写入指定文件')
    parser.add_argument('--nogui', action='store_true', help='无界面模式（仅命令行）')
    args = parser.parse_args()

    if args.nogui or args.input or args.diskc:
        # headless / CLI 模式
        root = tk.Tk()
        root.withdraw()
        app = TranslationApp(root)
        app.headless = True
        # 设置输出目录
        if args.output:
            app.output_dir = args.output
        # 准备日志
        if args.log:
            app._ensure_logger(args.log)
        else:
            app._ensure_logger()

        # 设置打包选项
        app.pack_var.set(bool(args.pack))

        # 设置语言列表
        if args.langs:
            langs = [s.strip() for s in args.langs.split(',') if s.strip()]
            app.selected_langs = langs
        else:
            app.selected_langs = app.default_langs.copy()

        # 准备输入文件列表
        if args.diskc:
            if os.path.isdir(args.diskc):
                app.log(f'DiskC工作流模式: {args.diskc}')
                app.setup_diskc_sources(args.diskc)
            else:
                app.log(f'DiskC路径无效: {args.diskc}')
                exit(1)
        elif args.input:
            if os.path.isdir(args.input) and args.mode == 'batch':
                app.source_folder = args.input
                app.refresh_file_list()
            elif os.path.isfile(args.input) and args.mode == 'single':
                # 单文件处理
                prepared = app.prepare_file(args.input)
                if prepared:
                    app.source_files = [prepared]
                else:
                    app.log('无法准备输入文件，退出')
                    exit(1)
            else:
                # 如果用户指定文件，但没有选择single模式，尝试作为单文件处理
                if os.path.isfile(args.input):
                    prepared = app.prepare_file(args.input)
                    if prepared:
                        app.source_files = [prepared]
                    else:
                        app.log('无法准备输入文件，退出')
                        exit(1)
                else:
                    app.log('输入路径无效')
                    exit(1)

        # 启动处理流程（同步等待完成）
        app.log('命令行模式: 开始处理')
        app.start_translation()

        # 等待翻译完成（轮询）
        try:
            while not app.translation_complete:
                time.sleep(0.5)
        except KeyboardInterrupt:
            app.log('用户中断')
            exit(2)

        # 生成XML
        app.generate_xml_files()
        app.save_translation_table()
        app.log('命令行模式: 处理完成')
        if app.logger:
            app.logger.info('处理完成')
        exit(0)
    else:
        root = tk.Tk()
        app = TranslationApp(root)
        root.mainloop()